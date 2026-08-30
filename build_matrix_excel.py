#!/usr/bin/env python3
"""
Aggregate the failure-stage matrix over four datasets x five backends
into memory_failure_matrix.xlsx.

Column groups follow the stage decomposition:
    Summary / Storage / Retrieval / Reasoning / Memory Performance
Each group is further split by dataset. A blank cell means either that the run
has not finished or that the cell does not exist structurally; see the
"Definitions" sheet.

Everything is read from each experiment's own scores.json and never recomputed,
so after changing a run it is enough to re-run this script.

    ./venv_memos/bin/python build_matrix_excel.py
"""

import os
import re
import json
from collections import Counter, defaultdict
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

BASE = os.path.dirname(os.path.abspath(__file__))
OUT = os.path.join(BASE, "memory_failure_matrix.xlsx")

# ── Which run corresponds to each backend in each of the four experiments ───
BACKENDS = [
    # (display name, halumem run, locomo run, longmemeval run, memfail run, LLM)
    ("Mem0 v1",   "mem0_oss-v1_31b_u2",              "mem0-v1_31b",      "mem0-v1_31b",      "results_5q_mem0v1",    "gemma-4-31B-it"),
    ("Mem0 v2",   "mem0_oss-v2_31b_u2",              "mem0-v2_31b",      "mem0-v2_31b",      "results_5q_mem0v2",    "gemma-4-31B-it"),
    ("StructMem", "structmem-user2nd_gemma431b_probe","structmem-sm_31b", "structmem-sm_31b", "results_5q_structmem", "gemma-4-31B-it"),
    ("A-MEM",     "amem-user2nd_gemma431b_probe",    "amem-amem_31b",    "amem-amem_31b",    "results_5q_amem",      "gemma-4-31B-it"),
    ("Letta",     "letta-user2nd_gemma431b_probe",   "letta-letta_31b",  "letta-letta_31b",  "results_5q_letta",     "openai-proxy/gemma-4-31B-it"),
    # Same runs, but P4 counts the message history visible at answering time
    # (recall memory) instead of only core blocks plus archival. Letta has three
    # memory tiers, and counting only the latter two underestimates what it can
    # actually see. Both rows are kept side by side: the gap between them is
    # itself a finding. No matching version was run for LongMemEval or MemFail,
    # so those cells stay empty.
    ("Letta (with history)", "letta-u2_ctx",                "letta-ctx_31b",    "",                 "",                     "openai-proxy/gemma-4-31B-it"),

    # ── Batch 2 (2026-08-19) ────────────────────────────────────────────────
    # Same backends and same sampling parameters; the only difference is that the
    # adapters now carry staged cost instrumentation (token_tracker's ingest / qa /
    # other buckets), which is why only this batch has Cost columns. It sits
    # alongside batch 1 rather than replacing it: the score gap between the two is
    # the run-to-run variance under identical parameters, which is itself a result.
    # HaluMem differs in one respect: these rows carry the users #3 and #4 sample,
    # which has 37 Dynamic Update questions against user #1's 4, so the
    # end-to-end evidence for post-update answering is far stronger here.
    ("Mem0 v1 · 0819",   "mem0_oss-v1_cost_u34",   "mem0-v1_cost",      "mem0-v1_cost",      "results_cost_mem0v1",    "gemma-4-31B-it"),
    ("Mem0 v2 · 0819",   "mem0_oss-v2_cost_u34",   "mem0-v2_cost",      "mem0-v2_cost",      "results_cost_mem0v2",    "gemma-4-31B-it"),
    # The 0819 StructMem run had an ingestion fault; per user decision it is
    # replaced by the ablation control E0. Scope caveats: E4B extraction model,
    # HaluMem user #1, LongMemEval limited to the 5 knowledge-update questions.
    ("StructMem · 0819", "structmem-ablate_e0_baseline", "structmem-ablate_e0_baseline", "structmem-ablate_e0_baseline", "results_ablate_e0_baseline", "gemma-4-E4B-it"),
    ("A-MEM · 0819",     "amem-amem_cost_u34",     "amem-amem_cost",    "amem-amem_cost",    "results_cost_amem",      "gemma-4-31B-it"),
    ("Letta · 0819",     "letta-letta_cost_u34",   "letta-letta_cost",  "letta-letta_cost",  "results_cost_letta",     "openai-proxy/gemma-4-31B-it"),

    # ── Batch 4 (2026-08-24): StructMem state ablation, M1 / M3 / M4 ────────
    # Five arms over a reduced but identical slice: LongMemEval knowledge-update
    # x5 and HaluMem user #1. Only the state-handling modules differ between
    # arms; embedding, candidate retrieval, update_queue, top-k, thresholds and
    # sampling are held fixed, so a gap between rows is attributable to the
    # module that was switched on.
    #   E0 original update/delete/ignore   E1 +non-destructive state bank
    #   E2 +summary sync                   E3 +state-aware retrieval/QA
    #   E4 all three
    # LoCoMo and MemFail are intentionally empty: this ablation was scoped to
    # the two benchmarks that actually exercise knowledge update.
    # All four benchmarks now, and note the LLM: the arms run gemma-4-E4B-it on
    # HaluMem user #1 (Martin_Mark), whereas the older StructMem rows above ran
    # gemma-4-31B-it on a different user. The batch-4 rows are therefore only
    # comparable to each other and to E0, never to the older rows.
    ("StructMem E0 baseline", "structmem-ablate_e0_baseline", "structmem-ablate_e0_baseline", "structmem-ablate_e0_baseline", "results_ablate_e0_baseline", "gemma-4-E4B-it"),
    ("StructMem E1 M1",       "structmem-ablate_e1_m1",       "structmem-ablate_e1_m1",       "structmem-ablate_e1_m1",       "results_ablate_e1_m1",       "gemma-4-E4B-it"),
    ("StructMem E2 M1+M3",    "structmem-ablate_e2_m1_m3",    "structmem-ablate_e2_m1_m3",    "structmem-ablate_e2_m1_m3",    "results_ablate_e2_m1_m3",    "gemma-4-E4B-it"),
    ("StructMem E3 M1+M4",    "structmem-ablate_e3_m1_m4",    "structmem-ablate_e3_m1_m4",    "structmem-ablate_e3_m1_m4",    "results_ablate_e3_m1_m4",    "gemma-4-E4B-it"),
    ("StructMem E4 full",     "structmem-ablate_e4_full",     "structmem-ablate_e4_full",     "structmem-ablate_e4_full",     "results_ablate_e4_full",     "gemma-4-E4B-it"),

    # ── Batch 5 (2026-08-29): the ablation redone on gemma-4-31B-it ─────────
    # Batch 4 ingested with gemma-4-E4B-it, the default in eval_structmem.py,
    # while every comparator (Mem0 v1/v2, A-MEM, Letta) uses gemma-4-31B-it on
    # LongMemEval, HaluMem and LoCoMo. Those arms therefore could not be placed
    # beside the other architectures at all. This batch re-runs them on 31B and
    # widens the slice:
    #   LongMemEval  21 questions (knowledge-update x6 plus 3 of each other
    #                type) rather than the 5 knowledge-update questions, so the
    #                column covers all six types. The 6 knowledge-update
    #                questions are a strict superset of batch 4's five.
    #   HaluMem      E0 runs users #3 and #4, the slice the four comparator
    #                architectures ran, so E0 is comparable across backends.
    #                E1 and E3 run user #3 only (half the ingest, inside E0's
    #                set): read the E1/E3-vs-E0 HaluMem delta as indicative,
    #                since it compares one user against two.
    #   LoCoMo       conv-26, 199 questions, extraction 31B and judge E4B,
    #                matching the comparator runs exactly.
    #   MemFail      points at the BATCH 4 directories on purpose. All five
    #                architectures ingested MemFail with gemma-4-E4B-it, so the
    #                E4B ablation runs are already the correct, comparable
    #                measurement; re-running them on 31B would put StructMem on
    #                a different model from every comparator. The MemFail cells
    #                in these rows are therefore E4B while the rest is 31B.
    ("StructMem E0 · 31B", "structmem-ablate_e0_baseline_31b", "structmem-ablate_e0_baseline_31b", "structmem-ablate_e0_baseline_31b", "results_ablate_e0_baseline", "gemma-4-31B-it"),
    ("StructMem E1 · 31B", "structmem-ablate_e1_m1_31b",       "structmem-ablate_e1_m1_31b",       "structmem-ablate_e1_m1_31b",       "results_ablate_e1_m1",       "gemma-4-31B-it"),
    ("StructMem E3 · 31B", "structmem-ablate_e3_m1_m4_31b",    "structmem-ablate_e3_m1_m4_31b",    "structmem-ablate_e3_m1_m4_31b",    "results_ablate_e3_m1_m4",    "gemma-4-31B-it"),

    # ── Batch 6 (2026-08-30): the five-architecture comparison, all on 31B ──
    # These rows point at runs that already appear above; they are repeated here
    # because no single existing batch is a valid cross-architecture comparison:
    #
    #   Batch 2 holds the four comparators on gemma-4-31B-it, but its StructMem
    #   row is gemma-4-E4B-it (and a much smaller slice: 5 LongMemEval questions
    #   and 164 HaluMem, against 22 and 360), so ranking StructMem inside batch 2
    #   compares a different model on different data.
    #
    #   Batch 5 has StructMem on 31B, but only StructMem: E0/E1/E3 are ablation
    #   arms of one architecture, with E0 as their shared baseline.
    #
    # Batch 6 is the intersection that is actually comparable: every row is
    # gemma-4-31B-it, LoCoMo is conv-26's 199 questions and MemFail the same 35
    # for all five. StructMem E0 keeps its batch 5 row as the ablation baseline;
    # duplicating it here is what lets one run serve both comparisons.
    #
    # Not perfectly matched: StructMem E0 answered 21 LongMemEval and 353 HaluMem
    # questions against the comparators' 22 and 360. Every metric here is a rate,
    # so the gap does not bias them, but the denominators are not identical.
    ("Mem0 v1 ⑥",     "mem0_oss-v1_cost_u34",   "mem0-v1_cost",     "mem0-v1_cost",     "results_cost_mem0v1",   "gemma-4-31B-it"),
    ("Mem0 v2 ⑥",     "mem0_oss-v2_cost_u34",   "mem0-v2_cost",     "mem0-v2_cost",     "results_cost_mem0v2",   "gemma-4-31B-it"),
    ("StructMem ⑥",   "structmem-ablate_e0_baseline_31b", "structmem-ablate_e0_baseline_31b", "structmem-ablate_e0_baseline_31b", "results_ablate_e0_baseline", "gemma-4-31B-it"),
    ("A-MEM ⑥",       "amem-amem_cost_u34",     "amem-amem_cost",   "amem-amem_cost",   "results_cost_amem",     "gemma-4-31B-it"),
    ("Letta ⑥",       "letta-letta_cost_u34",   "letta-letta_cost", "letta-letta_cost", "results_cost_letta",    "openai-proxy/gemma-4-31B-it"),
]

# Which batch each row belongs to. Batch 2 uses exactly the same sampling as
# batch 1 (LoCoMo conv-26, the first 3 questions of each LongMemEval type, and
# HaluMem's 2nd user), so the two batches can be compared directly.
BATCH = {"Mem0 v1": "① 08-14", "Mem0 v2": "① 08-14", "StructMem": "① 08-14",
         "A-MEM": "① 08-14", "Letta": "① 08-14", "Letta (with history)": "① 08-14",
         "Mem0 v1 · 0819": "② 08-19", "Mem0 v2 · 0819": "② 08-19",
         "StructMem · 0819": "② 08-19", "A-MEM · 0819": "② 08-19",
         "Letta · 0819": "② 08-19",
         "StructMem E0 baseline": "④ ablation", "StructMem E1 M1": "④ ablation",
         "StructMem E2 M1+M3": "④ ablation", "StructMem E3 M1+M4": "④ ablation",
         "StructMem E4 full": "④ ablation",
         "StructMem E0 · 31B": "⑤ 31B", "StructMem E1 · 31B": "⑤ 31B",
         "StructMem E3 · 31B": "⑤ 31B",
         "Mem0 v1 ⑥": "⑥ 31B compare", "Mem0 v2 ⑥": "⑥ 31B compare",
         "StructMem ⑥": "⑥ 31B compare", "A-MEM ⑥": "⑥ 31B compare",
         "Letta ⑥": "⑥ 31B compare"}

FRAME = {"mem0_oss": "mem0_oss", "mem0": "mem0", "structmem": "structmem",
         "amem": "amem", "letta": "letta"}


# Ingestion granularity per backend per dataset: whether one add call covers a
# single turn or a whole session. This is an adapter choice, not an architectural
# property, and the same backend can differ across datasets (mem0 is turn-level on
# LoCoMo but session-level on LongMemEval and HaluMem). It directly affects both
# memory volume and extraction quality, so it must be read alongside any /turn ratio.
GRANULARITY = {
    #  backend      LoCoMo     LongMemEval  HaluMem    MemFail
    "Mem0 v1":   ("turn",    "session",   "session", "turn"),
    "Mem0 v2":   ("turn",    "session",   "session", "turn"),
    "StructMem": ("session", "session",   "session", "turn"),
    "A-MEM":     ("turn",    "turn",      "turn",    "turn"),
    "Letta":     ("session", "session",   "session", "turn"),
    "Letta (with history)": ("session", "session", "session", "turn"),
    "Mem0 v1 · 0819":   ("turn",    "session",   "session", "turn"),
    "Mem0 v2 · 0819":   ("turn",    "session",   "session", "turn"),
    "StructMem · 0819": ("session", "session",   "session", "turn"),
    "A-MEM · 0819":     ("turn",    "turn",      "turn",    "turn"),
    "Letta · 0819":     ("session", "session",   "session", "turn"),
    "Mem0 v1 · u34":   ("turn",    "session",   "session", "turn"),
    "Mem0 v2 · u34":   ("turn",    "session",   "session", "turn"),
    "StructMem · u34": ("session", "session",   "session", "turn"),
    "A-MEM · u34":     ("turn",    "turn",      "turn",    "turn"),
    "Letta · u34":     ("session", "session",   "session", "turn"),
    # Batch 4 is StructMem throughout, so granularity is unchanged across arms.
    "StructMem E0 baseline": ("session", "session", "session", "turn"),
    "StructMem E1 M1":       ("session", "session", "session", "turn"),
    "StructMem E2 M1+M3":    ("session", "session", "session", "turn"),
    "StructMem E3 M1+M4":    ("session", "session", "session", "turn"),
    "StructMem E4 full":     ("session", "session", "session", "turn"),
    # Batch 5 is the same StructMem configuration, only the LLM and slice differ.
    "Mem0 v1 ⑥":        ("turn",    "session",   "session", "turn"),
    "Mem0 v2 ⑥":        ("turn",    "session",   "session", "turn"),
    "StructMem ⑥":      ("session", "session",   "session", "turn"),
    "A-MEM ⑥":          ("turn",    "turn",      "turn",    "turn"),
    "Letta ⑥":          ("session", "session",   "session", "turn"),
    "StructMem E0 · 31B":    ("session", "session", "session", "turn"),
    "StructMem E1 · 31B":    ("session", "session", "session", "turn"),
    "StructMem E3 · 31B":    ("session", "session", "session", "turn"),
}


# Turns actually fed into the memory systems in this round, used to normalize
# memory volume. Calibration: A-MEM writes exactly one entry per turn, so its
# entries-per-turn should come out to exactly 1.00. It does on all four datasets,
# which confirms these turn counts are right.
TURNS = {
    "locomo":  419,     # conv-26:19 sessions
    "halumem": 3242,    # user #1 / 77 sessions (batches 1 and 2)
    "longmem": 485,     # median haystack turns over the 18 questions (each builds its own store)
    "memfail": 45,      # fact sentences in the storage phase across the 35 questions
}

# Batch 3 runs HaluMem users #3 and #4, totalling 3,210 + 2,960 = 6,170 messages,
# unlike user #1's 3,242. Reusing the old denominator would inflate entries-per-turn
# by a factor of 1.9.
TURNS_BY_BATCH = {
    # Batch 2 now carries the HaluMem users #3 and #4 sample, so its HaluMem
    # denominators are those of that sample, not user #1.
    "② 08-19": {"halumem": 6170},
    # Batch 4 runs HaluMem user #1 = Martin_Mark (2,806 messages over 65 content
    # sessions), not the Johnson_Joseph user the earlier batches used, and its
    # LongMemEval slice is the 5 knowledge-update questions (median haystack 468
    # turns) rather than the 18-question stratified sample. Reusing the old
    # denominators would misstate entries-per-turn for both.
    "④ ablation": {"halumem": 2806, "longmem": 468},
    # Batch 5: HaluMem E0 is users #3 and #4 (6,170 messages, the comparator
    # slice); LongMemEval is the 21-question sample whose median haystack is 485
    # turns. E1 and E3 run one HaluMem user, so their entries-per-turn on that
    # dataset is computed against a denominator twice their true size and reads
    # low; the raw entry counts beside it are the honest figure.
    "⑤ 31B": {"halumem": 6170, "longmem": 485},
    # Batch 6: every row is the HaluMem users #3 and #4 sample, so it takes batch
    # 2's HaluMem denominator. LongMemEval is the one place the rows differ: the
    # four comparators ran the 22-question sample and StructMem the 21-question
    # one, whose median haystack is 485 rather than the default. Only entries-per-
    # turn is denominated that way, so that single cell in the StructMem row is
    # computed on a slightly different base; every rate is unaffected.
    "⑥ 31B compare": {"halumem": 6170},
}


def turns_for(batch: str, dataset: str) -> int:
    return TURNS_BY_BATCH.get(batch, {}).get(dataset, TURNS[dataset])


# Entry counts and HaluMem per-question-type accuracy can only be computed from
# the per-question raw output, and those files run to tens of megabytes each, so
# they are not tracked in git. scripts/export_derived_cache.py extracts the computed
# values into a cache of a few kilobytes: when the raw files are present this
# recomputes as before, and when they are absent (a plain clone, for instance) it
# reads the cache. Both paths yield the same numbers.
_CACHE_PATH = os.path.join(BASE, "results_derived_cache.json")
try:
    with open(_CACHE_PATH, encoding="utf-8") as _f:
        DERIVED_CACHE = json.load(_f)
except Exception:
    DERIVED_CACHE = {"store_size": {}, "halumem_per_type": {}}


def store_size(dataset: str, run: str):
    """How many entries this run stored on this dataset. Returns (total, per-question median).

    LongMemEval builds a separate store per question, so the total is a sum over
    18 questions and only the per-question median is a comparable scale.
    """
    import statistics
    try:
        if dataset == "locomo":
            frame = run.split("-")[0]
            f = os.path.join(BASE, "locomo_experiment", "results", run,
                             f"{frame}_locomo_results.jsonl")
            d = json.loads(open(f, encoding="utf-8").readline()).get("memory_dump")
            return (len(d), None) if d else (None, None)
        if dataset == "longmem":
            frame = run.split("-")[0]
            f = os.path.join(BASE, "longmemeval_experiment", "results", run,
                             f"{frame}_lme_results.jsonl")
            ns = [len(json.loads(l).get("all_memories") or [])
                  for l in open(f, encoding="utf-8") if l.strip()]
            return (sum(ns), statistics.median(ns)) if ns else (None, None)
        if dataset == "halumem":
            # Use store_after from the last session (the full store at that point)
            # rather than summing each session's extracted_memories:
            #   - Letta's extracted_memories is a cumulative snapshot, so summing
            #     would double-count by an order of magnitude
            #   - mem0 issues UPDATE/DELETE, so summing additions also overstates
            #     the final store size
            frame = "mem0_oss" if run.startswith("mem0_oss") else run.split("-")[0]
            f = os.path.join(BASE, "halumem_experiment", "results", run,
                             f"{frame}_eval_results.jsonl")
            us = [json.loads(l) for l in open(f, encoding="utf-8") if l.strip()]
            total = 0
            for u in us:
                last = None
                for sess in u["sessions"]:
                    if sess.get("store_after") is not None:
                        last = sess["store_after"]
                if last is not None:
                    total += len(last)
                else:   # only backends without store_after fall back to summing
                    total += sum(len(sess.get("extracted_memories") or [])
                                 for sess in u["sessions"])
            return (total, None)
        if dataset == "memfail":
            return (memfail(run).get("mf_store"), None)
    except Exception:
        pass
    cached = DERIVED_CACHE.get("store_size", {}).get(f"{dataset}/{run}")
    if cached:
        return tuple(cached)
    return (None, None)


def _load(path):
    try:
        with open(path, encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return None


def halumem(run):
    d = _load(os.path.join(BASE, "halumem_experiment", "results", run, "mem0_oss_scores.json"))
    if d is None:
        frame = run.split("-")[0]
        d = _load(os.path.join(BASE, "halumem_experiment", "results", run, f"{frame}_scores.json"))
    if not d:
        return {}
    mu = d.get("memory_update", {})
    qa = d.get("question_answering", {})
    at = d.get("qa_attribution") or {}
    mi = d.get("memory_integrity", {})
    ma = d.get("memory_accuracy", {})
    return {
        "hal_f1":        d.get("memory_extraction_f1"),
        "hal_integrity": mi.get("recall(all)"),
        "hal_target":    ma.get("target_accuracy(all)"),
        "hal_interf":    ma.get("interference_accuracy(all)"),
        "hal_update":    mu.get("correct_update_memory_ratio(all)"),
        "hal_upd_omis":  mu.get("omission_update_memory_ratio(all)"),
        "hal_upd_hall":  mu.get("hallucination_update_memory_ratio(all)"),
        "hal_p4":        at.get("retrieval_ratio"),
        "hal_p5":        at.get("generation_ratio"),
        "hal_storage":   at.get("storage_ratio"),
        "hal_unknown":   at.get("unknown_ratio"),
        "hal_qa":        qa.get("correct_qa_ratio(all)"),
        "hal_qa_hall":   qa.get("hallucination_qa_ratio(all)"),
        "hal_qa_omis":   qa.get("omission_qa_ratio(all)"),
        "hal_n":         qa.get("qa_num"),
        # The four update categories are denominated over all update points, but
        # only some are successfully adjudicated, so the (all) figures sum to less
        # than 1. The (valid) figures give the distribution over adjudicated ones.
        "hal_upd_valid":   mu.get("correct_update_memory_ratio(valid)"),
        "hal_upd_judged":  (mu.get("update_memory_valid_num") / mu["update_memory_num"]
                            if mu.get("update_memory_num") else None),
        "hal_upd_n":       mu.get("update_memory_num"),
    }


def locomo(run):
    frame = run.split("-")[0]
    d = _load(os.path.join(BASE, "locomo_experiment", "results", run, f"{frame}_locomo_scores.json"))
    if not d:
        return {}
    e = d.get("extraction") or {}
    p = d.get("probe") or {}
    r = d.get("retrieval") or {}
    if "skipped" in e:
        e = {}
    return {
        "loc_p1":      e.get("memory_integrity_recall"),
        "loc_acc":     e.get("memory_accuracy_precision"),
        "loc_f1":      e.get("memory_extraction_f1"),
        "loc_spk":     e.get("speaker_confusion_ratio"),
        "loc_recall":  r.get("recall@5"),
        "loc_ndcg":    r.get("ndcg@5"),
        "loc_p4":      p.get("P4_sufficient"),
        "loc_p1scope": p.get("P1_scoped_sufficient"),
        "loc_p5":      p.get("P5_fail_given_P4"),
        "loc_p5_n":    p.get("P5_n"),
        "loc_qa":      d.get("qa_accuracy_all"),
        "loc_tf1":     d.get("token_f1_all"),
        "loc_n":       d.get("qa_num"),
        "loc_attr_sum":  (p.get("attribution") or {}).get("SUMMARY"),
        "loc_attr_retr": (p.get("attribution") or {}).get("RETRIEVAL"),
        "loc_attr_reas": (p.get("attribution") or {}).get("REASONING"),
    }


def longmem(run):
    frame = run.split("-")[0]
    d = _load(os.path.join(BASE, "longmemeval_experiment", "results", run, f"{frame}_lme_scores.json"))
    if not d:
        return {}
    p = d.get("probe") or {}
    r = d.get("retrieval") or {}
    per = d.get("per_type") or {}
    ku = (per.get("knowledge-update") or {}).get("accuracy")
    return {
        "lme_p1":     p.get("P1_sufficient"),
        "lme_p4":     p.get("P4_sufficient"),
        "lme_p5":     p.get("P5_fail_given_P4"),
        "lme_p5_n":   p.get("P5_n"),
        "lme_ku":     ku,
        "lme_recall": r.get("recall@5"),
        "lme_ndcg":   r.get("ndcg@5"),
        "lme_qa":     d.get("qa_accuracy_all"),
        "lme_n":      d.get("qa_num"),
        "lme_attr_sum":  (p.get("attribution") or {}).get("SUMMARY"),
        "lme_attr_retr": (p.get("attribution") or {}).get("RETRIEVAL"),
        "lme_attr_reas": (p.get("attribution") or {}).get("REASONING"),
    }


# ── Cost and latency ───────────────────────────────────────────────────────
# Source is each run's {frame}_token_usage.json, measured in place by
# halumem_experiment/token_tracker.py while the experiment runs (not estimated
# afterwards). It sorts LLM calls into three buckets:
#     ingest  memory writes (extraction and update as each turn/session is fed in)
#     qa      answering (retrieval + generation)
#     other   judge / probes / warm-up, explicitly excluded, so the cost figures
#             carry no evaluation overhead
#
# Runs executed before 2026-08-18 recorded a single total_tokens with no
# bucketing. That legacy format always returns None and the table shows
# "legacy format" as a reminder that a re-run is needed to get figures.
_COST_FILE = {
    "longmem": ("longmemeval_experiment", "{frame}_lme"),
    "locomo":  ("locomo_experiment",      "{frame}_locomo"),
    "halumem": ("halumem_experiment",     "{frame}"),
}


def cost(dataset: str, run: str, prefix: str):
    if not run:
        return {}
    try:
        if dataset == "memfail":
            # MemFail 的 run 名就是 memfail_experiment/ 底下的目錄名,五個子集各有
            # 一份 token_usage。必須用 run 限定路徑:先前用 glob 抓「最新的一份」,
            # 導致三個批次全讀到同一個檔,第一批(舊格式、根本沒量成本)也被填上值。
            import glob
            hits = glob.glob(os.path.join(BASE, "memfail_experiment", run, "**",
                                          "memfail_token_usage.json"), recursive=True)
            if not hits:
                return {}
            # 五個子集加總後再算每單位,單看一個子集不能代表整個 MemFail
            tot = {"ingest": {"calls": 0, "total_tokens": 0, "units": 0, "secs": []},
                   "qa":     {"calls": 0, "total_tokens": 0, "units": 0, "secs": []}}
            for h in hits:
                dd = _load(h) or {}
                for ph in ("ingest", "qa"):
                    b = dd.get(ph) or {}
                    if not isinstance(b, dict):
                        continue
                    tot[ph]["calls"] += b.get("calls") or 0
                    tot[ph]["total_tokens"] += b.get("total_tokens") or 0
                    tot[ph]["units"] += b.get("units") or 0
                    if b.get("sec_per_unit_median") is not None:
                        tot[ph]["secs"].append(b["sec_per_unit_median"])
            def _per(ph, key):
                n = tot[ph]["units"]
                return round(tot[ph][key] / n, 3) if n else None
            def _sec(ph):
                ss = tot[ph]["secs"]
                return round(sum(ss) / len(ss), 3) if ss else None
            return {
                f"{prefix}_in_calls": _per("ingest", "calls"),
                f"{prefix}_in_tok":   _per("ingest", "total_tokens"),
                f"{prefix}_in_sec":   _sec("ingest"),
                f"{prefix}_qa_calls": _per("qa", "calls"),
                f"{prefix}_qa_tok":   _per("qa", "total_tokens"),
                f"{prefix}_qa_sec":   _sec("qa"),
            }
        # 其餘三個資料集:每個 run 一份 token_usage
        sub, pat = _COST_FILE[dataset]
        frame = ("mem0_oss" if run.startswith("mem0_oss") else run.split("-")[0])
        path = os.path.join(BASE, sub, "results", run,
                            pat.format(frame=frame) + "_token_usage.json")
        d = _load(path)
        if not d:
            return {}
        ing, qa = d.get("ingest"), d.get("qa")
        if not isinstance(ing, dict) or not isinstance(qa, dict):
            return {f"{prefix}_cost_old": "legacy format"}       # legacy run, no buckets
        return {
            f"{prefix}_in_calls": ing.get("calls_per_unit"),
            f"{prefix}_in_tok":   ing.get("tokens_per_unit"),
            f"{prefix}_in_sec":   ing.get("sec_per_unit_median"),
            f"{prefix}_qa_calls": qa.get("calls_per_unit"),
            f"{prefix}_qa_tok":   qa.get("tokens_per_unit"),
            f"{prefix}_qa_sec":   qa.get("sec_per_unit_median"),
        }
    except Exception:
        return {}


def memfail(run):
    """Read this run's TOTAL row from the experiment_results.md that compare_runs writes.

    Column order: Run | Dataset | Q# | Correct | Acc | Storage | Summary | Retr | Reason | Store
    """
    md = os.path.join(BASE, "memfail_experiment", "experiment_results.md")
    try:
        for line in open(md, encoding="utf-8"):
            c = [x.strip() for x in line.strip().strip("|").split("|")]
            if len(c) >= 10 and c[0] == run and c[1] == "TOTAL":
                n = int(c[2])
                return {
                    "mf_n":       n,
                    "mf_correct": float(c[4]),
                    "mf_storage": int(c[5]) / n if n else None,
                    "mf_summary": int(c[6]) / n if n else None,
                    "mf_retr":    int(c[7]) / n if n else None,
                    "mf_reason":  int(c[8]) / n if n else None,
                    "mf_store":   int(c[9]),
                }
    except Exception:
        pass
    return {}


def memfail_subsets(run):
    """The same file's non-TOTAL rows, one entry per subset.

    MemFail is the one dataset whose official output is already per subset, so
    nothing has to be recomputed from per-question files here: the four stage
    error counts are read straight out of the table and divided by that subset's
    own question count.
    """
    md = os.path.join(BASE, "memfail_experiment", "experiment_results.md")
    out = {}
    try:
        for line in open(md, encoding="utf-8"):
            c = [x.strip() for x in line.strip().strip("|").split("|")]
            if len(c) < 10 or c[0] != run or c[1] in ("TOTAL", "Dataset"):
                continue
            n = int(c[2])
            if not n:
                continue
            out[c[1]] = {
                "mf_n":       n,
                "mf_correct": float(c[4]),
                "mf_storage": int(c[5]) / n,
                "mf_summary": int(c[6]) / n,
                "mf_retr":    int(c[7]) / n,
                "mf_reason":  int(c[8]) / n,
                "mf_store":   int(c[9]),
            }
    except Exception:
        pass
    return out


# ── Stage failure rates (definition adopted 2026-08-25) ─────────────────────
# Denominator: all adjudicated questions. Numerator: questions attributed to
# that stage AND answered wrong. A stage failure that still produced the right
# answer is not counted as a failure; it is reported separately as "lucky".
# NO_WRITE is folded into SUMMARY. The three rates sum to the error rate, which
# is 1 minus accuracy, and share their definition with MemFail's official
# summary_error / retr_error / reason_error.
_SF_STAGE = {"SUMMARY": "p1", "NO_WRITE": "p1", "RETRIEVAL": "p4", "REASONING": "p5"}
_SF_NEUTRAL = {"OK"}
_SF_ABSTAIN = {"P5b_FAIL", "P5b_OK"}
_SF_EXCLUDE = {"UNKNOWN", "NO_DUMP", "UNADJUDICATED", None}


def _sf_from_detail(path, prefix):
    """Compute the stage failure rates from a per-question probe detail file."""
    if not os.path.exists(path):
        return {}
    fail = {"p1": 0, "p4": 0, "p5": 0}
    lucky = n = 0
    with open(path, encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            r = json.loads(line)
            v = r.get("verdict")
            if v in _SF_EXCLUDE or v in _SF_ABSTAIN:
                continue
            n += 1
            st = _SF_STAGE.get(v)
            if st is None:
                continue
            if r.get("is_correct") is True:
                lucky += 1
            else:
                fail[st] += 1
    if not n:
        return {}
    out = {f"{prefix}_sf_{k}": round(v / n, 4) for k, v in fail.items()}
    out[f"{prefix}_sf_err"] = round(sum(fail.values()) / n, 4)
    out[f"{prefix}_sf_lucky"] = lucky
    out[f"{prefix}_sf_n"] = n
    return out


def stage_rates(dataset, run, prefix):
    if not run:
        return {}
    frame = run.split("-")[0]
    if dataset == "longmem":
        return _sf_from_detail(os.path.join(
            BASE, "longmemeval_experiment", "results", run,
            f"{frame}_lme_probe_detail.jsonl"), prefix)
    if dataset == "locomo":
        return _sf_from_detail(os.path.join(
            BASE, "locomo_experiment", "results", run,
            f"{frame}_locomo_probe_detail.jsonl"), prefix)
    if dataset == "halumem":
        # HaluMem uses the unified probe: P1 and P4 both judged with the same
        # sufficiency prompt as the other two benchmarks.
        d = _load(os.path.join(BASE, "halumem_experiment", "results", run,
                               f"{frame}_probe_unified_scores.json"))
        if not d:
            return {}
        return {
            f"{prefix}_sf_p1":    d.get("P1_fail"),
            f"{prefix}_sf_p4":    d.get("P4_fail"),
            f"{prefix}_sf_p5":    d.get("P5_fail"),
            f"{prefix}_sf_err":   d.get("attributable_error_rate"),
            f"{prefix}_sf_lucky": d.get("not_counted"),
            f"{prefix}_sf_n":     d.get("n"),
        }
    return {}


# ── Column definitions: (group, header, key, source description) ────────────
COLUMNS = [
    ("",          "Run name",   "run_name", "Row label in this table"),
    ("",          "Backend",    "backend",  "Memory architecture"),
    ("",          "LLM",        "llm",      "Extraction LLM (judge is always gemma-4-E4B-it)"),
    ("",          "Batch",       "batch",    "(1) = original batch, 2026-08-14; (2) = re-run batch, 2026-08-19 (only this one has cost columns)"),

    # Within every group the dataset order is fixed: LongMemEval, LoCoMo,
    # HaluMem, MemFail. Header markers: ↑ higher is better, ↓ lower is better.
    # Differences in denominators are documented in the Definitions sheet.
    ("Summary",   "LongMemEval P1 fail (all) ↓", "lme_sf_p1", "Stage failure rate (definition adopted 2026-08-25): share of all adjudicated questions attributed to this stage AND answered wrong. A stage failure that still produced the right answer is not counted; it is reported under Attribution summary. Same definition as MemFail's official error metrics. SUMMARY includes NO_WRITE."),
    ("Summary",   "LoCoMo extraction recall ↑", "loc_p1",
     "extraction_locomo.py; observations treated as golden memories, strict=2. "
     "NOTE: the denominator is the number of golden observations (184 for conv-26), "
     "not the number of questions, so this is NOT the same quantity as "
     "LoCoMo P1 fail (all) and cannot be broken down by subset."),
    ("Summary",   "LoCoMo precision ↑",    "loc_acc", "Precision counterpart of the above (whether what was recorded is correct)"),
    ("Summary",   "LoCoMo F1 ↑",           "loc_f1",  "Harmonic mean of the two above"),
    ("Summary",   "LoCoMo P1 fail (all) ↓", "loc_sf_p1", "Stage failure rate (definition adopted 2026-08-25): share of all adjudicated questions attributed to this stage AND answered wrong. A stage failure that still produced the right answer is not counted; it is reported under Attribution summary. Same definition as MemFail's official error metrics."),
    ("Summary",   "HaluMem recall ↑",      "hal_integrity", "Official memory_integrity recall (all)"),
    ("Summary",   "HaluMem precision ↑",   "hal_target",    "Official target_accuracy (all); this is the precision used by f1"),
    ("Summary",   "HaluMem F1 ↑",          "hal_f1",  "Official memory_extraction_f1"),
    ("Summary",   "HaluMem interference ↑","hal_interf",
     "Official interference_accuracy (all). The dataset seeds 125 decoy memory points "
     "(memory_source=interference) that read plausibly but never happened. This metric "
     "scores the OPPOSITE way round to recall: a decoy earns a point only when the judge "
     "finds it ABSENT from the store, so it measures correct rejection. It belongs to "
     "Summary because it is still about what extraction did, only the suppression side "
     "rather than the retention side. Two caveats: it is NOT part of memory_extraction_f1 "
     "(which uses recall and target_accuracy only), and it is in tension with recall, "
     "since an indiscriminate extractor scores high on recall and low here. Read the two "
     "together; either one alone is misleading."),
    ("Summary",   "HaluMem P1 fail (all) ↓", "hal_sf_p1", "probe_halumem_unified.py; scope taken from the evidence's source session. Stage failure rate (definition adopted 2026-08-25): share of all adjudicated questions attributed to this stage AND answered wrong. A stage failure that still produced the right answer is not counted; it is reported under Attribution summary. Same definition as MemFail's official error metrics."),
    ("Summary",   "MemFail summary_error ↓","mf_summary","Official analyze_errors summary_error as a share of all questions"),

    ("Storage",   "HaluMem update ↑",      "hal_update", "Official correct_update_memory_ratio (all); omission and hallucination are its subcategories and are not listed separately"),
    ("Storage",   "MemFail storage_error ↓","mf_storage","Official analyze_errors not_stored as a share of all questions"),

    ("Retrieval", "LongMemEval P4 fail (all) ↓", "lme_sf_p4", "Stage failure rate (definition adopted 2026-08-25): share of all adjudicated questions attributed to this stage AND answered wrong. A stage failure that still produced the right answer is not counted; it is reported under Attribution summary. Same definition as MemFail's official error metrics."),
    ("Retrieval", "LongMemEval Recall@5 ↑","lme_recall","Official retrieval metric (session level)"),
    ("Retrieval", "LongMemEval NDCG@5 ↑",  "lme_ndcg",  "Official retrieval metric (session level)"),
    ("Retrieval", "LoCoMo P4 fail (all) ↓", "loc_sf_p4", "Stage failure rate (definition adopted 2026-08-25): share of all adjudicated questions attributed to this stage AND answered wrong. A stage failure that still produced the right answer is not counted; it is reported under Attribution summary. Same definition as MemFail's official error metrics."),
    ("Retrieval", "LoCoMo Recall@5 ↑",     "loc_recall","Official retrieval metric (turn level, exact dia_id match)"),
    ("Retrieval", "LoCoMo NDCG@5 ↑",       "loc_ndcg",  "Official retrieval metric (turn level)"),
    ("Retrieval", "HaluMem P4 fail (all) ↓", "hal_sf_p4", "probe_halumem_unified.py; re-judged with the shared sufficiency prompt. Stage failure rate (definition adopted 2026-08-25): share of all adjudicated questions attributed to this stage AND answered wrong. A stage failure that still produced the right answer is not counted; it is reported under Attribution summary. Same definition as MemFail's official error metrics."),
    ("Retrieval", "MemFail retr_error ↓",  "mf_retr",   "Official analyze_errors not_retrieved as a share of all questions"),

    # P5 = share of questions that passed P4 (the answer really was in the
    # retrieved context) yet were still answered wrong; the denominator is the
    # P4-passing questions. The definition is the same across all four datasets.
    # Denominator sizes vary a lot (LongMemEval has only 2 to 14), so read these
    # against the question counts in the Scale group. Per-metric denominators are
    # documented in the Definitions sheet.
    ("Reasoning", "LongMemEval P5 fail (all) ↓", "lme_sf_p5", "Stage failure rate (definition adopted 2026-08-25): share of all adjudicated questions attributed to this stage AND answered wrong. A stage failure that still produced the right answer is not counted; it is reported under Attribution summary. Same definition as MemFail's official error metrics."),
    ("Reasoning", "LoCoMo P5 fail (all) ↓", "loc_sf_p5", "Stage failure rate (definition adopted 2026-08-25): share of all adjudicated questions attributed to this stage AND answered wrong. A stage failure that still produced the right answer is not counted; it is reported under Attribution summary. Same definition as MemFail's official error metrics."),
    ("Reasoning", "HaluMem P5 fail (all) ↓", "hal_sf_p5", "Stage failure rate (definition adopted 2026-08-25): share of all adjudicated questions attributed to this stage AND answered wrong. A stage failure that still produced the right answer is not counted; it is reported under Attribution summary. Same definition as MemFail's official error metrics."),
    ("Reasoning", "MemFail reason_error ↓","mf_reason",  "Official analyze_errors reasoning_error as a share of all questions"),

    # Totals that are not stage-specific: the three stage rates sum to the error
    # rate, and "lucky" counts the questions where a stage failed yet the answer
    # came out right (not counted as failures).
    ("Attribution summary", "LongMemEval error ↓", "lme_sf_err",   "P1 + P4 + P5; equals 1 minus accuracy"),
    ("Attribution summary", "LongMemEval lucky n", "lme_sf_lucky", "Stage failed but answered correctly; excluded from the numerators"),
    ("Attribution summary", "LoCoMo error ↓",      "loc_sf_err",   "Same definition"),
    ("Attribution summary", "LoCoMo lucky n",      "loc_sf_lucky", "Same definition"),
    ("Attribution summary", "HaluMem error ↓",     "hal_sf_err",   "Same definition"),
    ("Attribution summary", "HaluMem lucky n",     "hal_sf_lucky", "Same definition; highest for Letta, which answers from conversation history"),

    ("Memory Performance", "LongMemEval QA ↑", "lme_qa", "Official LLM-judge accuracy"),
    ("Memory Performance", "LoCoMo QA ↑",      "loc_qa", "Official LLM-judge accuracy"),
    ("Memory Performance", "LoCoMo token F1 ↑","loc_tf1","Metric from the original LoCoMo paper"),
    ("Memory Performance", "HaluMem QA ↑",     "hal_qa", "Official correct_qa_ratio (all)"),
    ("Memory Performance", "MemFail correct ↑","mf_correct","Official overall accuracy"),

    ("Memory volume", "LME granularity",      "lme_gran",      "Whether one add call covers a turn or a whole session"),
    ("Memory volume", "LME entries/question",   "lme_store_med", "Each question builds its own store; median over the 18 questions"),
    ("Memory volume", "LME /turn",       "lme_store_pt",  "Median entry count / 485 (median haystack turns)"),
    ("Memory volume", "LME entries total",  "lme_store",     "Sum over the 18 per-question stores (scale reference only, not comparable across backends)"),
    ("Memory volume", "LoCoMo granularity",   "loc_gran",      "Whether one add call covers a turn or a whole session"),
    ("Memory volume", "LoCoMo entries",   "loc_store",     "Entries written for conv-26 (419 turns)"),
    ("Memory volume", "LoCoMo /turn",    "loc_store_pt",  "Entries / 419"),
    ("Memory volume", "HaluMem granularity",  "hal_gran",      "Whether one add call covers a turn or a whole session"),
    ("Memory volume", "HaluMem entries",  "hal_store",     "Total entries written for 1 user / 77 sessions (3,242 turns)"),
    ("Memory volume", "HaluMem /turn",   "hal_store_pt",  "Entries / 3,242"),
    ("Memory volume", "MemFail granularity",  "mf_gran",       "Whether one add call covers a turn or a session (fixed by the benchmark harness)"),
    ("Memory volume", "MemFail entries",  "mf_store",      "Store column from the official compare_runs output"),
    ("Memory volume", "MemFail /turn",   "mf_store_pt",   "Entries / 45"),

    # Cost: measured in place by token_tracker.py while the experiment runs, with
    # judge and probe overhead excluded. A "unit" is the span covered by one add
    # call (turn or session, see the granularity columns), so calls/unit must be
    # read together with granularity: A-MEM's 1.0 is per turn, StructMem's is per
    # session.
    ("Cost", "LME ingest calls/unit",  "lme_in_calls", "LLM calls needed to ingest one unit"),
    ("Cost", "LME ingest tok/unit",    "lme_in_tok",   "Tokens to ingest one unit (prompt + completion)"),
    ("Cost", "LME ingest sec/unit",     "lme_in_sec",   "End-to-end seconds to ingest one unit (median)"),
    ("Cost", "LME answer calls/question",    "lme_qa_calls", "LLM calls to answer one question (retrieval side included)"),
    ("Cost", "LME answer tok/question",      "lme_qa_tok",   "Tokens to answer one question"),
    ("Cost", "LME answer sec/question",       "lme_qa_sec",   "End-to-end seconds to answer one question (median)"),
    ("Cost", "LoCoMo ingest calls/unit","loc_in_calls", "As above"),
    ("Cost", "LoCoMo ingest tok/unit",  "loc_in_tok",   "As above"),
    ("Cost", "LoCoMo ingest sec/unit",   "loc_in_sec",   "As above"),
    ("Cost", "LoCoMo answer calls/question",  "loc_qa_calls", "As above"),
    ("Cost", "LoCoMo answer tok/question",    "loc_qa_tok",   "As above"),
    ("Cost", "LoCoMo answer sec/question",     "loc_qa_sec",   "As above"),
    ("Cost", "HaluMem ingest calls/unit","hal_in_calls","As above"),
    ("Cost", "HaluMem ingest tok/unit",  "hal_in_tok",  "As above"),
    ("Cost", "HaluMem ingest sec/unit",   "hal_in_sec",  "As above"),
    ("Cost", "HaluMem answer calls/question",  "hal_qa_calls","As above"),
    ("Cost", "HaluMem answer tok/question",    "hal_qa_tok",  "As above"),
    ("Cost", "HaluMem answer sec/question",     "hal_qa_sec",  "As above"),
    ("Cost", "MemFail ingest calls/unit","mf_in_calls", "MemFail is always turn level"),
    ("Cost", "MemFail ingest tok/unit",  "mf_in_tok",   "As above"),
    ("Cost", "MemFail ingest sec/unit",   "mf_in_sec",   "As above"),
    ("Cost", "MemFail answer calls/question",  "mf_qa_calls", "As above"),
    ("Cost", "MemFail answer tok/question",    "mf_qa_tok",   "As above"),
    ("Cost", "MemFail answer sec/question",     "mf_qa_sec",   "As above"),

    ("Scale", "LME questions",     "lme_n",     "3 questions per type"),
    ("Scale", "LME haystack", "lme_scale", "Haystack size per question"),
    ("Scale", "LoCoMo",      "loc_scale", "conversation / session / turn"),
    ("Scale", "LoCoMo questions",  "loc_n",     "All QA in that conversation"),
    ("Scale", "HaluMem",     "hal_scale", "user / session / turn"),
    ("Scale", "HaluMem questions", "hal_n",     "All QA for that user"),
    ("Scale", "HaluMem update points","hal_upd_n","Number of updated memory points"),
    ("Scale", "MemFail questions", "mf_n",      "Question count after sampling the 5 subsets"),

    ("Quality", "LoCoMo speaker confusion ↓","loc_spk","Fact recorded correctly but attributed to the wrong speaker (specific to LoCoMo two-party dialogue)"),
    ("Quality", "HaluMem update judged rate ↑","hal_upd_judged","Successfully judged / all update points; below 1 means update is underestimated"),
    ("Quality", "HaluMem attribution unknown ↓","hal_unknown","Share that qa_attribution could not adjudicate"),
]

# ── Stage failure rates lead their own stage group ──────────────────────────
# P1 / P4 / P5 are this study's headline numbers; the official per-dataset
# metrics that follow them are supporting evidence. The reordering happens here
# rather than in the list above so that list can stay grouped by dataset, which
# is the order it is easiest to edit in. Sorting is stable and keyed on each
# group's first position, so the stage groups themselves do not move.
_STAGE_FIRST = {
    "Summary":   ("lme_sf_p1", "loc_sf_p1", "hal_sf_p1", "mf_summary"),
    "Retrieval": ("lme_sf_p4", "loc_sf_p4", "hal_sf_p4", "mf_retr"),
    "Reasoning": ("lme_sf_p5", "loc_sf_p5", "hal_sf_p5", "mf_reason"),
}
_GRP_POS = {}
for _i, _c in enumerate(COLUMNS):
    _GRP_POS.setdefault(_c[0], _i)


def _stage_first_key(c):
    lead = _STAGE_FIRST.get(c[0], ())
    return (_GRP_POS[c[0]], 0, lead.index(c[2])) if c[2] in lead else (_GRP_POS[c[0]], 1, 0)


COLUMNS.sort(key=_stage_first_key)


# what kind of memory a correct answer demands, not the shape of the answer.
PURPOSE_GROUPS = [
    ("1",  "Single-point recall",
     "One memory entry suffices; the answer is a value held directly in it."),
    ("2a", "Multi-hop composition",
     "The answer is reached by traversing entries in order: the next one cannot be "
     "located until the previous has been read, because it is found through a bridge "
     "the previous supplies. A missing link breaks the traversal and the question is "
     "unanswerable, not merely incomplete. This is multi-hop in the sense the QA "
     "literature uses it. Whether a benchmark constructs the chain deliberately "
     "(MemFail's long_hop verifies strict transitivity) or simply labels a subset "
     "multi-hop (LoCoMo, HaluMem) is a property of the benchmark, not of the memory "
     "demand, so both sit here; the difference is recorded in the run notes instead. "
     "Note that MemFail's long_hop is 5-way multiple choice while the other two are "
     "open-ended, so their scores are not directly comparable."),
    ("2b", "Multi-memory, parallel",
     "Several entries must all be present, but every one of them is reachable from "
     "the original query: no traversal is needed, only completeness. Missing one "
     "leaves the answer incomplete or, where the entries qualify one another, "
     "confidently wrong. The typical failure is extraction merging them into a "
     "single superordinate concept, or deduplication treating them as one entry so "
     "that they overwrite each other. MemFail's conditional_hard belongs here: its "
     "behavior, condition and linking sentences all describe the same entity and "
     "surface together on one query, so the demand is to keep all three rather than "
     "to hop between them."),
    ("3",  "Temporal reasoning",
     "A correct answer requires the memory to retain timestamps (date differences, "
     "ordering, most recent occurrence). Dropping time information during "
     "extraction guarantees failure."),
    ("4",  "Post-update value",
     "The same fact has been updated and the latest value must be returned. This "
     "sits at the Storage stage: whether the stale value was removed or superseded."),
    ("5",  "Abstention and correction",
     "The correct behavior is not to answer but to admit the information is absent "
     "from memory, or to point out that the question rests on a false premise. The "
     "scoring logic differs from the other groups: it judges whether the system "
     "correctly avoided answering, not whether it produced some value, so scores in "
     "this group should not be pooled with the others."),
    ("6",  "Application and extrapolation",
     "No memory entry holds the answer verbatim. What is remembered has to be "
     "applied to a situation that was never discussed, or extrapolated past what "
     "was actually stated, so a faithful recall of the right entry is necessary "
     "but not sufficient. This separates a reasoning demand from a retrieval "
     "demand: a system can pass P4 here, having surfaced exactly the right "
     "memories, and still answer wrong."),
]

# The single source of truth for the subset -> purpose group mapping.
_SUBSETS = [
    # (key prefix, subset name, display name, purpose group code)
    ("lme_sub_", "single-session-user",          "single-session-user",      "1"),
    ("lme_sub_", "single-session-assistant",     "single-session-assistant", "1"),
    ("lme_sub_", "single-session-preference",    "single-session-preference","1"),
    ("lme_sub_", "multi-session",                "multi-session",            "2b"),
    ("lme_sub_", "temporal-reasoning",           "temporal-reasoning",       "3"),
    ("lme_sub_", "knowledge-update",             "knowledge-update",         "4"),
    ("loc_sub_", "single_hop",                   "cat4 single_hop",          "1"),
    ("loc_sub_", "multi_hop",                    "cat1 multi_hop",           "2a"),
    ("loc_sub_", "open_domain",                  "cat3 open_domain",         "6"),
    ("loc_sub_", "temporal",                     "cat2 temporal",            "3"),
    ("loc_sub_", "adversarial",                  "cat5 adversarial",         "5"),
    ("hal_sub_", "Basic Fact Recall",            "Basic Fact Recall",        "1"),
    ("hal_sub_", "Multi-hop Inference",          "Multi-hop Inference",      "2a"),
    ("hal_sub_", "Generalization & Application", "Generalization & App.",    "6"),
    ("hal_sub_", "Dynamic Update",               "Dynamic Update",           "4"),
    ("hal_sub_", "Memory Boundary",              "Memory Boundary",          "5"),
    ("hal_sub_", "Memory Conflict",              "Memory Conflict",          "5"),
    ("mf_sub_",  "persona_retrieval",            "persona_retrieval",        "1"),
    ("mf_sub_",  "conditional_easy",             "conditional_easy",         "1"),
    ("mf_sub_",  "long_hop",                     "long_hop",                 "2a"),
    ("mf_sub_",  "conditional_hard",             "conditional_hard",         "2b"),
    ("mf_sub_",  "coexisting_facts",             "coexisting_facts",         "2b"),
]

# Keyed by the raw subset name and by the display name alike: LoCoMo's breakdown
# arrives already mapped through LOCOMO_CAT ("cat4 single_hop"), while the other
# three datasets hand back the raw question_type, and both must find their group.
_GROUP_NAME = {_code: _name for _code, _name, _desc in PURPOSE_GROUPS}
_SUBSET_GROUP = {}
_SUBSET_DISP = {}
for _p, _k, _disp, _g in _SUBSETS:
    # The Category cell spells the purpose group out; the short codes stay
    # internal, where they keep the table above readable.
    _SUBSET_GROUP[_k] = _SUBSET_GROUP[_disp] = _GROUP_NAME[_g]
    _SUBSET_DISP[_k] = _SUBSET_DISP[_disp] = _disp



# ── Per-subset notes for the Subset Map sheet ───────────────────────────────
# Only properties that change how a number should be read go here: what the
# subset demands, and anything that makes its score not comparable with its
# neighbours. Keyed by (prefix, raw subset name) so the two spellings of a
# LoCoMo subset cannot drift apart.
_SUBSET_NOTE = {
    ("lme_sub_", "single-session-user"):
        "Answer stated by the user inside one session.",
    ("lme_sub_", "single-session-assistant"):
        "Answer stated by the assistant inside one session. Extraction that keeps "
        "only the user side loses this subset entirely.",
    ("lme_sub_", "single-session-preference"):
        "A preference stated once, to be returned verbatim.",
    ("lme_sub_", "multi-session"):
        "Evidence spans several sessions. The official label does not say how the "
        "pieces relate, and they are reachable from the query without traversal.",
    ("lme_sub_", "temporal-reasoning"):
        "Requires timestamps to survive extraction.",
    ("lme_sub_", "knowledge-update"):
        "The same fact was updated; the latest value must be returned.",
    ("loc_sub_", "single_hop"):
        "One turn holds the answer.",
    ("loc_sub_", "multi_hop"):
        "Official multi-hop label; open-ended answer, so not directly comparable "
        "with MemFail's long_hop, which is multiple choice.",
    ("loc_sub_", "open_domain"):
        "Answer is not stated outright anywhere and must be extrapolated from what "
        "was remembered.",
    ("loc_sub_", "temporal"):
        "Timestamps must survive extraction. The sharpest single separator in the "
        "study: event-anchored storage answers it, rewrite-style extraction does not.",
    ("loc_sub_", "adversarial"):
        "The question rests on a false premise; the correct behavior is to say so.",
    ("hal_sub_", "Basic Fact Recall"):
        "One memory point holds the answer.",
    ("hal_sub_", "Multi-hop Inference"):
        "Open-ended composition across memory points.",
    ("hal_sub_", "Generalization & Application"):
        "The stored fact must be applied to a situation never discussed.",
    ("hal_sub_", "Dynamic Update"):
        "The latest value of an updated fact. Small subset; read as a pattern.",
    ("hal_sub_", "Memory Boundary"):
        "Abstention: the answer is deliberately absent from the record and the "
        "correct behavior is to say so. The official annotation therefore carries no "
        "evidence, the sufficiency judgment has nothing to apply to, and these "
        "questions are excluded from the stage-rate denominator. Accuracy is the "
        "only figure this subset reports; its P1, P4 and P5 cells read n/a by "
        "construction, not for want of a measurement.",
    ("hal_sub_", "Memory Conflict"):
        "Two values that were each correct at some point both sit in the store; the "
        "current one must be chosen. Unlike Memory Boundary it does carry evidence, "
        "so it has full stage rates.",
    ("mf_sub_",  "persona_retrieval"):
        "Half the graded queries name a distractor and require abstention, so this "
        "subset mixes a single-point recall demand with an abstention demand and "
        "cannot be split into the two groups.",
    ("mf_sub_",  "conditional_easy"):
        "The whole rule sits in one sentence; copying it verbatim succeeds.",
    ("mf_sub_",  "long_hop"):
        "A strictly transitive chain, verified by construction. Graded as 5-way "
        "multiple choice with shape-matched distractors, so its scores run high and "
        "do not compare with the open-ended multi-hop subsets.",
    ("mf_sub_",  "conditional_hard"):
        "The rule is split across three non-adjacent sentences (behavior, condition, "
        "link) describing one entity, so all three surface on a single query and the "
        "demand is to keep them, not to hop between them.",
    ("mf_sub_",  "coexisting_facts"):
        "N compatible preferences delivered in separate conversations; all must be "
        "returned. Graded all-or-nothing per question over five questions, so one "
        "missed item moves the subset score by 0.2.",
}


GROUP_FILL = {
    "":                    "F2F5F8",
    "Summary":             "E8F3F0",
    "Storage":             "E8EEF9",
    "Retrieval":           "F8F0DE",
    "Reasoning":           "F5E8F0",
    "Attribution summary": "EFEFEF",
    "Memory Performance":  "E6E6E6",
    "Memory volume":               "E8F0FA",
    "Cost":                 "FDF2E9",
    "Subset scores":             "EDF3EA",
    "Scale":                 "F0F2F5",
    "Quality":                 "FBEEE6",
}
GROUP_FONT = {
    "Summary": "16685A", "Storage": "2A4A8E", "Retrieval": "87590C",
    "Reasoning": "782D58", "Memory Performance": "333333", "": "141E27",
    "Memory volume": "1A4F8A", "Scale": "5C6B7A", "Quality": "8A5A2B", "Cost": "9C4221", "Subset scores": "3D6B2E",
}


def run_done(dataset: str, run: str) -> bool:
    """Whether this run's scores.json exists. Absent means the run has not finished
    (queued or in progress). An empty run name means the combination has no
    corresponding run at all; that counts as complete and renders as a dash.
    """
    if not run:
        return True
    if dataset == "halumem":
        rd = os.path.join(BASE, "halumem_experiment", "results", run)
        frame = "mem0_oss" if run.startswith("mem0_oss") else run.split("-")[0]
        return os.path.exists(os.path.join(rd, f"{frame}_scores.json"))
    if dataset == "locomo":
        frame = run.split("-")[0]
        return os.path.exists(os.path.join(BASE, "locomo_experiment", "results", run,
                                           f"{frame}_locomo_scores.json"))
    if dataset == "longmem":
        frame = run.split("-")[0]
        return os.path.exists(os.path.join(BASE, "longmemeval_experiment", "results", run,
                                           f"{frame}_lme_scores.json"))
    if dataset == "memfail":
        return bool(memfail(run))
    return False


# Column-key prefix to dataset, used to tell a blank cell that is "still running"
# from one that does not exist structurally
KEY_DATASET = {"hal_": "halumem", "loc_": "locomo", "lme_": "longmem", "mf_": "memfail"}


def collect():
    rows = []
    for name, hal, loc, lme, mf, llm in BACKENDS:
        # The label was hard-coded to gemma4-31B, which silently mislabelled the
        # batch-4 ablation rows: those run gemma-4-E4B-it. Derive it from the
        # row's own LLM instead.
        short_llm = llm.replace("openai-proxy/", "").replace("gemma-4-", "gemma4-").replace("-it", "")
        r = {"run_name": f"{name} · {short_llm}", "backend": name, "llm": llm,
             "batch": BATCH.get(name, "")}
        r.update(halumem(hal)); r.update(locomo(loc))
        r.update(longmem(lme)); r.update(memfail(mf))
        r["_done"] = {"halumem": run_done("halumem", hal), "locomo": run_done("locomo", loc),
                      "longmem": run_done("longmem", lme), "memfail": run_done("memfail", mf)}
        # Store size: necessary context for reading recall and precision, since
        # writing more entries makes high recall easier by construction
        r["loc_store"], _                  = store_size("locomo", loc)
        r["lme_store"], r["lme_store_med"] = store_size("longmem", lme)
        r["hal_store"], _                  = store_size("halumem", hal)
        r["mf_store"], _                   = store_size("memfail", mf)
        # Entries per turn: dataset scales differ, so only this ratio is
        # comparable across datasets
        for k, ds in [("loc_store", "locomo"), ("hal_store", "halumem"),
                      ("mf_store", "memfail")]:
            v = r.get(k)
            r[k + "_pt"] = (v / turns_for(r["batch"], ds)) if isinstance(v, (int, float)) else None
        r["lme_store_pt"] = (r["lme_store_med"] / TURNS["longmem"]
                             if isinstance(r.get("lme_store_med"), (int, float)) else None)
        # Cost and latency: only runs executed after 2026-08-18 carry staged data
        r.update(stage_rates("longmem", lme, "lme"))
        r.update(stage_rates("locomo", loc, "loc"))
        r.update(stage_rates("halumem", hal, "hal"))
        r.update(cost("longmem", lme, "lme")); r.update(cost("locomo", loc, "loc"))
        r.update(cost("halumem", hal, "hal")); r.update(cost("memfail", mf, "mf"))
        g = GRANULARITY.get(name, ("?",) * 4)
        r["loc_gran"], r["lme_gran"], r["hal_gran"], r["mf_gran"] = g
        # Scale: question counts alone do not convey workload, so user / session /
        # turn counts are added
        # Batch 5 splits by arm: E0 ran the comparator slice (users #3 and #4)
        # so it can sit beside the other architectures, while E1 and E3 ran only
        # user #3 to halve the ingest.
        if r["batch"] == "⑤ 31B":
            r["hal_scale"] = ("2 users (#3, #4) / 6,170 turns"
                              if "E0" in name
                              else "1 user (#3) / 3,210 turns")
        elif r["batch"] == "② 08-19":
            r["hal_scale"] = "2 users (#3, #4) / 6,170 turns"
        else:
            r["hal_scale"] = "1 user / 77 sessions / 3,242 turns"
        r["loc_scale"] = "1 conv (conv-26) / 19 sessions / 419 turns"
        # Batch 5 widened LongMemEval from 5 knowledge-update questions to a
        # 21-question sample spanning all six types.
        r["lme_scale"] = ("21 questions (KU x6 + 3 each of five types), median 485 turns"
                          if r["batch"] == "⑤ 31B"
                          else "about 50 sessions, 491 turns per question")
        rows.append(r)
    return rows


# Key prefix to dataset, used to split out the per-dataset sheets
KEY_PREFIX = {"lme_": "LongMemEval", "loc_": "LoCoMo", "hal_": "HaluMem", "mf_": "MemFail"}
SHEET_ORDER = ["LongMemEval", "LoCoMo", "HaluMem", "MemFail"]


# A second copy of the subset -> purpose group mapping used to sit here and was
# never read by anything. It is gone: _SUBSETS above is the only copy, so a
# reclassification cannot silently apply to one table and not the other.

# The experiments name their verdicts differently; normalize to five categories
# before comparing
VERDICT_MAP = {"OK": "OK", "SUMMARY": "SUMMARY", "NO_WRITE": "STORAGE",
               "RETRIEVAL": "RETRIEVAL", "NOT_RETRIEVED": "RETRIEVAL",
               "REASONING": "REASONING"}
VERDICT_COLS = ["OK", "SUMMARY", "STORAGE", "RETRIEVAL", "REASONING", "Other"]

LOCOMO_CAT = {"1": "multi_hop", "2": "temporal", "3": "open_domain",
              "4": "single_hop", "5": "adversarial"}


def _norm_verdicts(counts):
    from collections import Counter
    o = Counter()
    for k, v in (counts or {}).items():
        o[VERDICT_MAP.get(k, "Other")] += v
    return o


def subset_scores(name, hal, loc, lme, mf):
    """Official metrics for this backend on every subset of the four datasets.

    Returns {column_key: score}.

    Official metrics only: QA accuracy for LongMemEval and HaluMem, accuracy plus
    token F1 for LoCoMo, accuracy for MemFail. Retrieval (recall@k), extraction
    (P1), and update metrics are stored as overall values only in all four
    experiments, with no subset-level breakdown available.
    """
    out = {}

    # ── LongMemEval ── official per_type.accuracy
    if lme:
        frame = lme.split("-")[0]
        d = _load(os.path.join(BASE, "longmemeval_experiment", "results", lme,
                               f"{frame}_lme_scores.json")) or {}
        for t, st in (d.get("per_type") or {}).items():
            out[f"lme_sub_{t}"] = st.get("accuracy")

    # ── LoCoMo ── official per_category carries both accuracy and token_f1
    if loc:
        frame = loc.split("-")[0]
        d = _load(os.path.join(BASE, "locomo_experiment", "results", loc,
                               f"{frame}_locomo_scores.json")) or {}
        for t, st in (d.get("per_category") or {}).items():
            out[f"loc_sub_{t}"] = st.get("accuracy")
            out[f"loc_subf1_{t}"] = st.get("token_f1")

    # ── HaluMem ── the official output gives only the overall correct_qa_ratio,
    #    so per-type figures are recomputed from the per-question file. The
    #    recomputed overall value has been verified to match the official one.
    if hal:
        from collections import Counter
        frame = "mem0_oss" if hal.startswith("mem0_oss") else hal.split("-")[0]
        tot, cor = Counter(), Counter()
        try:
            path = os.path.join(BASE, "halumem_experiment", "results", hal,
                                f"{frame}_eval_detail.jsonl")
            with open(path, encoding="utf-8") as f:
                for line in f:
                    if not line.strip():
                        continue
                    r = json.loads(line)
                    if "result_type" not in r or "question" not in r:
                        continue
                    tot[r.get("question_type")] += 1
                    if r.get("result_type") == "Correct":
                        cor[r.get("question_type")] += 1
        except Exception:
            pass
        if tot:
            for t in tot:
                out[f"hal_sub_{t}"] = cor[t] / tot[t] if tot[t] else None
        else:   # per-question detail absent (a plain clone) -> read the cache
            for t, acc in (DERIVED_CACHE.get("halumem_per_type", {}).get(hal) or {}).items():
                out[f"hal_sub_{t}"] = acc

    # ── MemFail ── the md table compare_runs writes, one row per subset
    if mf:
        try:
            md = os.path.join(BASE, "memfail_experiment", "experiment_results.md")
            for line in open(md, encoding="utf-8"):
                c = [x.strip() for x in line.strip().strip("|").split("|")]
                if len(c) >= 10 and c[0] == mf and c[1] != "TOTAL":
                    out[f"mf_sub_{c[1]}"] = float(c[4])
        except Exception:
            pass
    return out

def dataset_of(key: str):
    """Which dataset a column key belongs to. The first three columns
    (run_name / backend / llm) belong to none and return None."""
    for pre, ds in KEY_PREFIX.items():
        if key.startswith(pre):
            return ds
    return None



# ── Per-subset breakdown for the dataset sheets ─────────────────────────────
# Every stage-failure rate uses the same denominator (adjudicated questions), so
# each one can simply be regrouped by subset. Recomputed from the per-question
# files; scores.json only ever stored the aggregate. Metrics whose denominator is
# NOT the question (extraction recall counts golden memories, memory volume and
# cost count the whole ingest) cannot be split and are marked "n/a".
LOCOMO_CAT = {"1": "cat1 multi_hop", "2": "cat2 temporal", "3": "cat3 open_domain",
              "4": "cat4 single_hop", "5": "cat5 adversarial"}

# Which verdict counts as which stage failure. NO_WRITE means the session wrote
# nothing at all, which is a summary failure.
# The unified HaluMem probe emits RETRIEVAL; the older probe_detail called the
# same thing NOT_RETRIEVED. Fold that alias into the shared stage map.
_SF_STAGE.setdefault("NOT_RETRIEVED", "p4")


def _read_jsonl(path):
    try:
        with open(path, encoding="utf-8") as f:
            return [json.loads(l) for l in f if l.strip()]
    except Exception:
        return []


def subset_breakdown(dataset: str, run: str):
    """{subset: {metric: value}} recomputed per subset, or {} when unavailable."""
    if not run:
        return {}
    if dataset == "locomo":
        frame = run.split("-")[0]
        rd = os.path.join(BASE, "locomo_experiment", "results", run)
        probe = _read_jsonl(os.path.join(rd, f"{frame}_locomo_probe_detail.jsonl"))
        detail = _read_jsonl(os.path.join(rd, f"{frame}_locomo_detail.jsonl"))
        keyof = lambda r: LOCOMO_CAT.get(str(r.get("category")), str(r.get("category")))
    elif dataset == "longmem":
        frame = run.split("-")[0]
        rd = os.path.join(BASE, "longmemeval_experiment", "results", run)
        probe = _read_jsonl(os.path.join(rd, f"{frame}_lme_probe_detail.jsonl"))
        detail = _read_jsonl(os.path.join(rd, f"{frame}_lme_detail.jsonl"))
        keyof = lambda r: r.get("question_type")
    elif dataset == "halumem":
        frame = "mem0_oss" if run.startswith("mem0_oss") else run.split("-")[0]
        rd = os.path.join(BASE, "halumem_experiment", "results", run)
        # probe_halumem_unified.py is the only HaluMem probe now: it carries P1 as
        # well as P4/P5 and emits the SUMMARY / RETRIEVAL / REASONING vocabulary of
        # Algorithm 1. The retired probe_halumem.py wrote *_probe_detail.jsonl with
        # neither P1 nor a SUMMARY verdict, and called retrieval failure
        # NOT_RETRIEVED; reading it silently produced P1 = P4 = 0, so there is
        # deliberately no fallback to it.
        probe = _read_jsonl(os.path.join(rd, f"{frame}_probe_unified.jsonl"))
        detail = [r for r in _read_jsonl(os.path.join(rd, f"{frame}_eval_detail.jsonl"))
                  if "result_type" in r and "question" in r]
        keyof = lambda r: r.get("question_type")
    else:
        return {}
    if not probe and not detail:
        return {}

    out = {}

    # Stage failure rates, from the probe verdicts
    by = defaultdict(list)
    for r in probe:
        by[keyof(r)].append(r)
    for sub, rs in by.items():
        # Abstention questions are excluded from the denominator: the right
        # behaviour there is to decline, so they never enter stage attribution.
        # This matches how the official stage-failure rates are computed.
        # Reuse the official definition verbatim so a subset row and the TOTAL row
        # are the same quantity: abstention and unadjudicable questions leave the
        # denominator, and a stage failure that still answered correctly (a lucky
        # hit) is not counted as a failure.
        adj = [r for r in rs if r.get("verdict") not in _SF_EXCLUDE
               and r.get("verdict") not in _SF_ABSTAIN]
        d = out.setdefault(sub, {})
        d["n"] = len(rs)              # every question, used for QA accuracy
        d["n_stage"] = len(adj)       # stage-failure denominator
        if not adj:
            continue                  # a purely abstention subset has no stage rates
        cnt = Counter()
        for r in adj:
            st = _SF_STAGE.get(r.get("verdict"))
            if st and r.get("is_correct") is not True:
                cnt[st] += 1
        for st in ("p1", "p4", "p5"):
            d[f"sf_{st}"] = cnt[st] / len(adj)

    # QA accuracy, token f1 and retrieval, from the evaluation detail
    by2 = defaultdict(list)
    for r in detail:
        by2[keyof(r)].append(r)
    for sub, rs in by2.items():
        d = out.setdefault(sub, {})
        d.setdefault("n", len(rs))
        ok = [r for r in rs if ("is_correct" in r or "result_type" in r)]
        if ok:
            d["qa"] = sum(1 for r in ok
                          if r.get("is_correct") or r.get("result_type") == "Correct") / len(ok)
        f1 = [r["token_f1"] for r in rs if isinstance(r.get("token_f1"), (int, float))]
        if f1:
            d["f1"] = sum(f1) / len(f1)
        for k in ("recall@5", "ndcg@5"):
            vals = [r["retrieval"][k] for r in rs
                    if isinstance(r.get("retrieval"), dict)
                    and isinstance(r["retrieval"].get(k), (int, float))]
            if vals:
                d[k.replace("@", "")] = sum(vals) / len(vals)
    return out

# Two columns that exist only on the dataset sheets, where a run occupies several
# rows. Run name stays identical to the Failure Matrix sheet on every one of a
# run's rows, so the same run is looked up the same way on both sheets; which row
# is which is carried by Sub-dataset instead.
_SUBSET_COLS = [
    ("", "Sub-dataset", "_sub_name",
     "The official subset this row scores, or TOTAL for the whole run. TOTAL is "
     "the value that appears on the Failure Matrix sheet."),
    ("", "Category", "_sub_group",
     "The purpose group this subset belongs to: what kind of memory a correct "
     "answer demands. The groups are defined at the bottom of this sheet. Two "
     "subsets sharing a group answer to the same demand and can be read against "
     "each other even across datasets."),
]


# ── Subset detail as extra columns, one row per run ─────────────────────────
# The sheet keeps the Failure Matrix's shape, one row per run, and widens each
# question-denominated metric into "TOTAL" plus one column per official subset.
# The subset columns sit at column outline level 1, so the 1/2 buttons above the
# sheet switch between TOTAL only and TOTAL plus every subset.
#
# Two earlier attempts are worth not repeating. Grouping subsets by purpose group
# kept four benchmarks in shared metric columns they do not share, so most cells
# were "n/a". Putting each subset on its own row fixed that but broke the thing
# the Failure Matrix is for: one row per run, every benchmark side by side.
# Widening the columns instead keeps both.
_SUB_DROP_GROUPS = {"Memory volume", "Scale", "Quality"}

# Column key -> the metric to read out of subset_breakdown / memfail_subsets.
# Anything absent here is a whole-run quantity (extraction recall is denominated
# over memory points, cost over an ingest batch) and stays a single column.
_SPLIT_METRIC = {
    "lme_sf_p1": "sf_p1", "lme_sf_p4": "sf_p4", "lme_sf_p5": "sf_p5",
    "lme_qa": "qa", "lme_recall": "recall5", "lme_ndcg": "ndcg5",
    "loc_sf_p1": "sf_p1", "loc_sf_p4": "sf_p4", "loc_sf_p5": "sf_p5",
    "loc_qa": "qa", "loc_tf1": "f1", "loc_recall": "recall5", "loc_ndcg": "ndcg5",
    "hal_sf_p1": "sf_p1", "hal_sf_p4": "sf_p4", "hal_sf_p5": "sf_p5", "hal_qa": "qa",
    "mf_summary": "mf_summary", "mf_storage": "mf_storage", "mf_retr": "mf_retr",
    "mf_reason": "mf_reason", "mf_correct": "mf_correct",
}
_DS_OF_PREFIX = {"lme_sub_": "longmem", "loc_sub_": "locomo",
                 "hal_sub_": "halumem", "mf_sub_": "memfail"}
# Declaration order, not per-run question counts, so a column sits in the same
# place on every row and the sheet can be read down a column.
_DS_SUBSETS = {ds: [(k, disp) for p, k, disp, g in _SUBSETS
                    if _DS_OF_PREFIX[p] == ds]
               for ds in ("longmem", "locomo", "halumem", "memfail")}
_KEY_DS = {"lme": "longmem", "loc": "locomo", "hal": "halumem", "mf": "memfail"}


def _split_title(title, label):
    """'LongMemEval P1 fail (all) \u2193' + 'temporal-reasoning' ->
    'LongMemEval P1 fail \u00b7 temporal-reasoning \u2193'"""
    arrow = title[-1] if title and title[-1] in "\u2191\u2193" else ""
    base = (title[:-1] if arrow else title).strip().replace(" (all)", "")
    return f"{base} \u00b7 {label} {arrow}".strip()


def subset_columns():
    """Failure Matrix columns, minus the un-splittable groups, with every
    splittable metric widened into TOTAL plus one column per subset.

    Returns (columns, levels): levels[i] is 1 for a subset column, 0 otherwise.
    """
    cols, levels = [], []
    for grp, title, key, desc in COLUMNS:
        if grp in _SUB_DROP_GROUPS:
            continue
        if key not in _SPLIT_METRIC:
            cols.append((grp, title, key, desc))
            levels.append(0)
            continue
        ds = _KEY_DS[key.split("_")[0]]
        cols.append((grp, _split_title(title, "TOTAL"), key,
                     desc + " | Whole-run value, identical to the Failure Matrix sheet."))
        levels.append(0)
        for raw, disp in _DS_SUBSETS[ds]:
            cols.append((grp, _split_title(title, disp), f"{key}@{raw}",
                         desc + f" | Restricted to the {disp} subset."))
            levels.append(1)
    return cols, levels


def _subset_cell(d, metric):
    """One subset cell: the value, "n/a", or None.

    "n/a" means the metric does not apply to this subset, and is only ever
    written when the measurement that would produce it did run. A stage rate is
    inapplicable when the subset has no adjudicated questions at all (HaluMem's
    Memory Boundary is entirely abstention, so its denominator is zero); token F1
    is inapplicable to LoCoMo's cat5, which has no answer column to score. If the
    probe has simply not been run for this subset yet, the cell stays None so it
    reads as a missing measurement, not as a structural absence.
    """
    if metric in d:
        return d[metric]
    if metric.startswith("sf_"):
        return "n/a" if "n_stage" in d else None    # n_stage present => probe ran
    return "n/a" if d else None                     # eval ran but no such metric


def subset_sheet_rows(rows):
    """The Failure Matrix rows, each carrying its per-subset values as well."""
    out = []
    for row, be in zip(rows, BACKENDS):
        r = dict(row)
        for ds in ("longmem", "locomo", "halumem", "memfail"):
            run = be[_DS_RUN_IDX[ds]]
            bd = (memfail_subsets(run) if ds == "memfail"
                  else subset_breakdown(ds, run)) if run else {}
            for key, metric in _SPLIT_METRIC.items():
                if _KEY_DS[key.split("_")[0]] != ds:
                    continue
                for raw, disp in _DS_SUBSETS[ds]:
                    # LoCoMo's breakdown is keyed by display name, the rest by the
                    # raw question_type, so both spellings are tried.
                    d = bd.get(raw) or bd.get(disp) or {}
                    r[f"{key}@{raw}"] = _subset_cell(d, metric)
        out.append(r)
    return out


def columns_for(ds: str):
    """Columns for a per-dataset sheet: the shared columns, Sub-dataset and
    Category, then that dataset's own, in the order they appear in COLUMNS."""
    cols = [c for c in COLUMNS if dataset_of(c[2]) in (None, ds)]
    at = next(i for i, c in enumerate(cols) if c[2] == "batch") + 1
    return cols[:at] + _SUBSET_COLS + cols[at:]



# ── Expand a dataset sheet into one row per subset ──────────────────────────
# Every stage-failure rate shares one denominator (adjudicated questions), so a
# subset row is just a regrouping. Anything counted per memory point (extraction
# recall) or per ingest batch (memory volume, cost) has no subset-level value and
# is written as "n/a" so it reads differently from a run that has not finished.
_SUBSET_KEYS = {
    "locomo":  {"qa": "loc_qa", "f1": "loc_tf1", "sf_p1": "loc_sf_p1",
                "sf_p4": "loc_sf_p4", "sf_p5": "loc_sf_p5",
                "recall5": "loc_recall", "ndcg5": "loc_ndcg"},
    "longmem": {"qa": "lme_qa", "sf_p1": "lme_sf_p1", "sf_p4": "lme_sf_p4",
                "sf_p5": "lme_sf_p5", "recall5": "lme_recall", "ndcg5": "lme_ndcg"},
    "halumem": {"qa": "hal_qa", "sf_p1": "hal_sf_p1", "sf_p4": "hal_sf_p4",
                "sf_p5": "hal_sf_p5"},
}
_DS_RUN_IDX = {"locomo": 2, "longmem": 3, "halumem": 1, "memfail": 4}


def expand_rows_for(ds: str, rows: list, columns: list):
    """Return rows for a dataset sheet: each run becomes its subsets plus TOTAL."""
    dskey = {"LoCoMo": "locomo", "LongMemEval": "longmem",
             "HaluMem": "halumem", "MemFail": "memfail"}[ds]
    owned = {c[2] for c in columns if dataset_of(c[2]) == ds}
    idx = _DS_RUN_IDX[dskey]
    out = []
    for row in rows:
        run = next((be[idx] for be in BACKENDS if be[0] == row["backend"]), "")
        if dskey == "memfail":
            # Already per subset upstream, so the values are copied over directly
            # rather than recomputed; sort by question count like the others.
            bd = memfail_subsets(run) if run else {}
            order = sorted(bd, key=lambda k: -bd[k]["mf_n"])
            fill = lambda sub: {k: v for k, v in bd[sub].items() if k in owned}
        else:
            bd = subset_breakdown(dskey, run) if run else {}
            order = sorted(bd, key=lambda k: -bd[k].get("n", 0))
            keymap = _SUBSET_KEYS[dskey]
            def fill(sub, _km=keymap):
                d = {c: bd[sub][m] for m, c in _km.items()
                     if c in owned and m in bd[sub]}
                n = bd[sub].get("n")
                for cand in ("loc_n", "lme_n", "hal_n"):
                    if cand in owned and n is not None:
                        d[cand] = n
                        break
                return d
        for sub in order:
            # Run name, backend, LLM and batch are repeated verbatim from the
            # Failure Matrix row so a run reads the same on both sheets.
            r = {k: row.get(k) for k in ("run_name", "backend", "llm", "batch")}
            r.update({"_sub_name": _SUBSET_DISP.get(sub, sub),
                      "_sub_group": _SUBSET_GROUP.get(sub, ""),
                      "_done": row.get("_done", {}), "_subset": True})
            for k in owned:
                r[k] = "n/a"                          # not measurable per subset
            r.update(fill(sub))
            out.append(r)
        total = dict(row)
        total["_sub_name"] = "TOTAL" if bd else ""
        out.append(total)
    return out


# ── Metric families: the same measurement under four dataset prefixes ───────
# On the long sheets a row belongs to exactly one dataset, so the four
# per-dataset copies of a metric are mutually exclusive and collapse into one
# column. Scanning P1 down a column across benchmarks is the whole point of
# those sheets, and four columns of which three are always blank defeats it.
#
# Two families need their synonyms reconciled first: LoCoMo calls its extraction
# recall "extraction recall" and HaluMem calls the same measurement "recall";
# MemFail calls its end-to-end accuracy "correct" where the other three say "QA".
#
# MemFail's summary_error, storage_error, retr_error and reason_error are NOT
# folded into P1/P4/P5. They come from the benchmark's own analyze_errors, not
# from this study's probes, and a column that silently mixed the two definitions
# is exactly the kind of thing that produces a wrong claim later.
_FAM_SYNONYM = {"recall ↑": "extraction recall ↑", "correct ↑": "QA ↑"}
_FAM_TITLE = {"P1 fail (all) ↓": "P1 fail ↓", "P4 fail (all) ↓": "P4 fail ↓",
              "P5 fail (all) ↓": "P5 fail ↓", "error ↓": "Error ↓",
              "lucky n": "Lucky n", "questions": "Questions",
              "entries": "Entries", "/turn": "Entries/turn",
              "granularity": "Granularity",
              "extraction recall ↑": "Extraction recall ↑",
              "precision ↑": "Extraction precision ↑",
              "F1 ↑": "Extraction F1 ↑"}
_DS_PREFIX = ("LongMemEval ", "LoCoMo ", "HaluMem ", "MemFail ", "LME ")


def _family(title):
    """The dataset-independent name of a metric, or None for a shared column."""
    for pre in _DS_PREFIX:
        if title.startswith(pre):
            rest = title[len(pre):].strip()
            return _FAM_SYNONYM.get(rest, rest) if rest else None
    return None


def merged_columns():
    """Failure Matrix columns with the three key columns inserted after Batch and
    every multi-dataset metric family collapsed to one column."""
    at = next(i for i, c in enumerate(COLUMNS) if c[2] == "batch") + 1
    head, tail = COLUMNS[:at], COLUMNS[at:]
    members = {}
    for grp, title, key, desc in tail:
        fam = _family(title)
        if fam:
            members.setdefault(fam, []).append(key)
    out, seen = [], set()
    for grp, title, key, desc in tail:
        fam = _family(title)
        if not fam or len(members[fam]) == 1:
            out.append((grp, title, key, desc))
            continue
        if fam in seen:
            continue
        seen.add(fam)
        out.append((grp, _FAM_TITLE.get(fam, fam), f"fam::{fam}",
                    desc + " Merged across " +
                    ", ".join(dataset_of(k) or "?" for k in members[fam]) +
                    "; a row carries the value for its own dataset."))
    return head + _LONG_KEY_COLS + out, members


def _fill_merged(dst, src, members):
    """Copy src's per-dataset values into their merged column keys."""
    for fam, keys in members.items():
        vals = [src.get(k) for k in keys if isinstance(src.get(k), (int, float))]
        dst[f"fam::{fam}"] = vals[0] if vals else "n/a"


# ── One row per run per dataset per subset ──────────────────────────────────
# Long format, replacing the earlier pooled-by-purpose-group shape. Each row is
# a single measurement: one run, one dataset, one official subset. Dataset,
# Sub-dataset and Category sit together right after Batch, so the three keys
# that identify a row read as one block.
#
# Nothing is pooled. The previous shape averaged the subsets a benchmark
# contributed to one group (HaluMem put both Memory Boundary and Memory Conflict
# into abstention), which forced a weight choice and hid the two components. Here
# both subsets keep their own row and the reader groups them if they want to.
#
# The metric columns are the Failure Matrix's, unchanged and complete. A row
# therefore carries values only in its own dataset's columns and n/a everywhere
# else, which is the price of stacking four benchmarks into one sheet and is
# paid deliberately: a column means exactly what it means on the Failure Matrix,
# with no renaming or remapping in between.
_LONG_DS_NAME = {"longmem": "LongMemEval", "locomo": "LoCoMo",
                 "halumem": "HaluMem", "memfail": "MemFail"}
_LONG_NCOL = {"longmem": "lme_n", "locomo": "loc_n", "halumem": "hal_n"}

_LONG_KEY_COLS = [
    ("", "Dataset", "_ds_name",
     "Which benchmark this row scores. Only this dataset's columns carry a value; "
     "the other three read n/a."),
    ("", "Sub-dataset", "_sub_name",
     "The official subset within that benchmark, or TOTAL for the whole run on it."),
    ("", "Category", "_sub_group",
     "The purpose group this subset belongs to: what kind of memory a correct "
     "answer demands. Subsets sharing a group answer to the same demand and can "
     "be read against each other across datasets. Rows whose Sub-dataset is TOTAL "
     "carry TOTAL here too, so a pivot excludes them with one filter."),
]


def long_columns():
    """Failure Matrix columns with Dataset, Sub-dataset and Category after Batch."""
    at = next(i for i, c in enumerate(COLUMNS) if c[2] == "batch") + 1
    return COLUMNS[:at] + _LONG_KEY_COLS + COLUMNS[at:]


def category_rows(rows, members):
    """One row per run per dataset per subset, plus that run's TOTAL per dataset."""
    owned = {c[2] for c in COLUMNS if dataset_of(c[2])}
    out = []
    for row, be in zip(rows, BACKENDS):
        head = {k: row.get(k) for k in ("run_name", "backend", "llm", "batch")}
        for ds in ("longmem", "locomo", "halumem", "memfail"):
            run = be[_DS_RUN_IDX[ds]]
            if not run:
                continue
            bd = (memfail_subsets(run) if ds == "memfail"
                  else subset_breakdown(ds, run)) or {}
            if not bd:
                continue
            dsname = _LONG_DS_NAME[ds]
            nkey = "mf_n" if ds == "memfail" else "n"
            for sub in sorted(bd, key=lambda k: -(bd[k].get(nkey) or 0)):
                r = dict(head)
                r.update({"_ds_name": dsname,
                          "_sub_name": _SUBSET_DISP.get(sub, sub),
                          "_sub_group": _SUBSET_GROUP.get(sub, ""),
                          "_done": row.get("_done", {}), "_subset": True})
                # Everything starts not-measurable: another dataset's column, or
                # this dataset's but denominated over something other than the
                # question (extraction recall, memory volume, cost).
                raw = {k: "n/a" for k in owned}
                d = bd[sub]
                if ds == "memfail":
                    for k, v in d.items():
                        if k in owned:
                            raw[k] = v
                else:
                    for metric, col in _SUBSET_KEYS[ds].items():
                        if col in owned and metric in d:
                            raw[col] = d[metric]
                    ncol = _LONG_NCOL[ds]
                    if ncol in owned and d.get("n") is not None:
                        raw[ncol] = d["n"]
                r.update(raw)
                _fill_merged(r, raw, members)
                out.append(r)
            # The run's TOTAL on this dataset, copied from the Failure Matrix row
            # so the two sheets cannot disagree. Here the whole-run metrics that
            # have no subset value (extraction recall, cost) are real again.
            tot = dict(head)
            tot.update({"_ds_name": dsname, "_sub_name": "TOTAL",
                        "_sub_group": "TOTAL", "_done": row.get("_done", {})})
            raw = {k: (row.get(k) if dataset_of(k) == dsname else "n/a")
                   for k in owned}
            tot.update(raw)
            _fill_merged(tot, raw, members)
            out.append(tot)
    return out


# ── The same rows, ordered so a purpose group reads as one block ────────────
# By Category is ordered by run, which answers "what did this run do". Sorting
# by group first answers the other question: "how does every architecture handle
# temporal reasoning", with LongMemEval and LoCoMo adjacent instead of pages
# apart. TOTAL rows sink to the bottom so they never break a group's block.
_BACKEND_ORDER = ["Mem0 v1", "Mem0 v2", "A-MEM", "Letta", "StructMem"]
_CODE_OF_GROUP = {name: code for code, name, _d in PURPOSE_GROUPS}


def _sort_key(r):
    grp = r.get("_sub_group") or ""
    code = _CODE_OF_GROUP.get(grp, "~")            # TOTAL and unknowns last
    be = str(r.get("backend") or "")
    fam = next((i for i, b in enumerate(_BACKEND_ORDER) if be.startswith(b)),
               len(_BACKEND_ORDER))
    ds = r.get("_ds_name") or ""
    dsi = SHEET_ORDER.index(ds) if ds in SHEET_ORDER else len(SHEET_ORDER)
    return (code, fam, be, dsi, str(r.get("_sub_name") or ""))


def group_sorted_rows(rows, members):
    return sorted(category_rows(rows, members), key=_sort_key)


def write_sheet(ws, columns, rows, col_levels=None):
    """Write one set of columns as a sheet: two header rows, merged group headers,
    best value in bold, and RUNNING markers."""
    thin = Side(style="thin", color="C8D0D8")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Two header rows: group, then column name
    for ci, (grp, title, key, _) in enumerate(columns, 1):
        g = ws.cell(row=1, column=ci, value=grp)
        g.fill = PatternFill("solid", fgColor=GROUP_FILL.get(grp, "FFFFFF"))
        g.font = Font(bold=True, size=10, color=GROUP_FONT.get(grp, "000000"))
        g.alignment = Alignment(horizontal="center", vertical="center")
        g.border = border

        h = ws.cell(row=2, column=ci, value=title)
        h.fill = PatternFill("solid", fgColor=GROUP_FILL.get(grp, "FFFFFF"))
        h.font = Font(bold=True, size=9, color=GROUP_FONT.get(grp, "000000"))
        h.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        h.border = border
        # Sub-dataset and Category hold spelled-out names far longer than their
        # own headers, so they are sized to the content instead of the title.
        ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = \
            {"_sub_ds": 14, "_sub_name": 26, "_sub_group": 30}.get(
                key, max(11, min(20, len(title) + 3)))

    # Merge the group header cells
    start = 1
    for ci in range(2, len(columns) + 2):
        cur = columns[ci - 1][0] if ci <= len(columns) else None
        prev = columns[start - 1][0]
        if cur != prev:
            if ci - 1 > start:
                ws.merge_cells(start_row=1, start_column=start, end_row=1, end_column=ci - 1)
            start = ci

    # Best value per column: max for ↑ columns, min for ↓ columns. Ties are all
    # bolded. Ranking only means something between rows that answered the same
    # questions, so each batch is its own comparison group. This used to test for
    # a "③" label that no row carries: the HaluMem users #3 and #4 sample is
    # batch 2, so every batch fell into one pool and batch 1 (user #1, 188
    # questions) was being ranked against batch 2 (users #3 and #4, 360 entirely
    # different questions).
    def _cmp_group(row):
        return str(row.get("batch", "")) or "main"

    # Only whole-run rows compete for the best value. A subset row is a slice of
    # one run, so letting it into the comparison would crown, say, the single
    # easiest category of one backend over every other backend's full score.
    ranked = [r for r in rows if not r.get("_subset")]
    best = {}
    for ci, (grp, title, key, _) in enumerate(columns, 1):
        if "↑" not in title and "↓" not in title:
            continue
        for g in {_cmp_group(r) for r in ranked}:
            vals = [row.get(key) for row in ranked
                    if _cmp_group(row) == g and isinstance(row.get(key), (int, float))]
            if vals:
                best[(ci, g)] = max(vals) if "↑" in title else min(vals)

    text_keys = {"run_name", "backend", "llm", "batch",
                 "_sub_ds", "_sub_name", "_sub_group"}
    for ri, row in enumerate(rows, 3):
        sub = bool(row.get("_subset"))
        for ci, (grp, title, key, _) in enumerate(columns, 1):
            v = row.get(key)
            c = ws.cell(row=ri, column=ci, value=v)
            c.border = border
            c.alignment = Alignment(horizontal="center", vertical="center")
            if isinstance(v, float):
                c.number_format = "0.0000"
            if key in text_keys:
                # Subset rows print their identity a shade lighter than the TOTAL
                # rows, enough to tell the two apart without making the subset
                # labels hard to read.
                c.font = (Font(size=9, color="3D444C") if sub
                          else Font(bold=(key in ("run_name", "backend")), size=9))
                c.alignment = Alignment(horizontal="left", vertical="center")
            elif (not sub and isinstance(v, (int, float))
                  and best.get((ci, _cmp_group(row))) == v):
                c.font = Font(bold=True, size=9, color="1A4F8A")
            if v is None:
                # A blank cell means one of two things: the run has not finished,
                # or the metric does not exist structurally (Letta has no turn
                # provenance, so LoCoMo recall@k simply cannot be computed)
                ds = {"LongMemEval": "longmem", "LoCoMo": "locomo",
                      "HaluMem": "halumem", "MemFail": "memfail"}.get(dataset_of(key))
                if ds and not row.get("_done", {}).get(ds, True):
                    c.value = "RUNNING"
                    c.font = Font(color="C05621", size=9, italic=True, bold=True)
                else:
                    c.value = "—"
                    c.font = Font(color="B0B8C0", size=9)

    # Outline the subset rows so the left-hand +/- collapses each run down to its
    # TOTAL. Level 1 is the subsets, level 2 (the outline's "2" button) shows them
    # all; summaryBelow says the TOTAL row sits underneath its own group, which is
    # how the rows are emitted. Groups start collapsed so a sheet opens on the
    # same view as the Failure Matrix.
    # Column outline: the subset columns collapse under the TOTAL that precedes
    # them, so summaryRight is False. The 1/2 buttons above the sheet then switch
    # between TOTAL only and TOTAL plus every subset.
    if col_levels and any(col_levels):
        for ci, lv in enumerate(col_levels, 1):
            if lv:
                cd = ws.column_dimensions[openpyxl.utils.get_column_letter(ci)]
                cd.outlineLevel = lv
                cd.hidden = True
        ws.sheet_properties.outlinePr.summaryRight = False
        ws.sheet_format.outlineLevelCol = max(col_levels)
        ws.sheet_view.showOutlineSymbols = True

    ws.sheet_properties.outlinePr.summaryBelow = True
    ws.sheet_properties.outlinePr.applyStyles = False
    grouped = False
    for ri, row in enumerate(rows, 3):
        if row.get("_subset"):
            ws.row_dimensions[ri].outlineLevel = 1
            ws.row_dimensions[ri].hidden = True
            grouped = True
    if grouped:
        ws.sheet_view.showOutlineSymbols = True
        # Excel sizes the outline gutter from this; without it the +/- buttons can
        # fail to appear even though the rows carry an outline level.
        ws.sheet_format.outlineLevelRow = 1

    # Freeze everything up to and including the identity columns
    last_text = max((ci for ci, c in enumerate(columns, 1)
                     if c[2] in ("run_name", "backend", "llm", "batch",
                                 "_sub_ds", "_sub_name", "_sub_group")), default=3)
    ws.freeze_panes = f"{openpyxl.utils.get_column_letter(last_text + 1)}3"
    ws.row_dimensions[1].height = 20
    ws.row_dimensions[2].height = 34


def _write_outline_level_col(path):
    """Add sheetFormatPr@outlineLevelCol, which openpyxl never serializes.

    openpyxl writes each column's own outlineLevel but derives no sheet-level
    maximum for columns the way it does for rows, and Excel sizes the outline
    gutter from that attribute: without it the collapse buttons above the columns
    can fail to appear even though the grouping is present. Rewritten in place
    over the saved file, touching only that one attribute.
    """
    import re
    import shutil
    import zipfile

    src = zipfile.ZipFile(path)
    tmp = path + ".tmp"
    with zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as dst:
        for item in src.infolist():
            data = src.read(item.filename)
            if item.filename.startswith("xl/worksheets/sheet"):
                x = data.decode("utf-8")
                lv = [int(m) for m in re.findall(r'<col [^>]*outlineLevel="(\d+)"', x)]
                if lv and "outlineLevelCol" not in x:
                    x = re.sub(r"<sheetFormatPr ",
                               f'<sheetFormatPr outlineLevelCol="{max(lv)}" ', x, count=1)
                    data = x.encode("utf-8")
            dst.writestr(item, data)
    src.close()
    shutil.move(tmp, path)


def build():
    rows = collect()
    wb = openpyxl.Workbook()

    # Master sheet: every column
    ws = wb.active
    ws.title = "Failure Matrix"
    write_sheet(ws, COLUMNS, rows)

    # The same comparison one level down: every subset of every benchmark
    sc, lv = subset_columns()
    write_sheet(wb.create_sheet("By Sub-dataset"), sc, subset_sheet_rows(rows), lv)

    # A third view: one row per run per dataset per subset, Category as a label
    lc, members = merged_columns()
    write_sheet(wb.create_sheet("By Category"), lc, category_rows(rows, members))

    # The same rows ordered group first, so one purpose group reads as a block
    write_sheet(wb.create_sheet("By Group"), lc, group_sorted_rows(rows, members))

    # One sheet per dataset: same structure, only that dataset's columns
    for ds in SHEET_ORDER:
        cols = columns_for(ds)
        write_sheet(wb.create_sheet(ds), cols, expand_rows_for(ds, rows, cols))

    # ── Definitions sheet ───────────────────────────────────────────────────
    ws2 = wb.create_sheet("Definitions")
    for ci, h in enumerate(["Stage", "Column", "Source / definition"], 1):
        c = ws2.cell(row=1, column=ci, value=h)
        c.font = Font(bold=True)
        c.fill = PatternFill("solid", fgColor="F2F5F8")
    for ri, (grp, title, key, desc) in enumerate(COLUMNS, 2):
        ws2.cell(row=ri, column=1, value=grp or "—")
        ws2.cell(row=ri, column=2, value=title)
        ws2.cell(row=ri, column=3, value=desc).alignment = Alignment(wrap_text=True)

    # ── What the [code] prefix on subset columns means ──────────────────────
    ri = len(COLUMNS) + 3
    t = ws2.cell(row=ri, column=1, value="Purpose group codes")
    t.font = Font(bold=True, size=11, color="3D6B2E")
    ws2.cell(row=ri, column=3,
             value="The Category column on each dataset sheet holds one of these codes, "
                   "naming the purpose group that row's subset belongs to. The grouping dimension is "
                   "what kind of memory a correct answer demands: how many entries, "
                   "whether timestamps are needed, whether the latest value is needed. "
                   "Whether external world knowledge is required, whether arithmetic is "
                   "involved, and whether the answer is open-ended are cross-cutting "
                   "attributes and do not define groups. Two subsets carrying the same "
                   "code answer to the same demand and can be read against each other "
                   "even when they come from different datasets."
             ).alignment = Alignment(wrap_text=True)
    for ci, h in enumerate(["Code", "Group", "Definition"], 1):
        c = ws2.cell(row=ri + 1, column=ci, value=h)
        c.font = Font(bold=True, size=10)
        c.fill = PatternFill("solid", fgColor="EDF3EA")
    for k, (code, gname, gdesc) in enumerate(PURPOSE_GROUPS, ri + 2):
        ws2.cell(row=k, column=1, value=code).font = Font(bold=True, size=10)
        ws2.cell(row=k, column=2, value=gname)
        ws2.cell(row=k, column=3, value=gdesc).alignment = Alignment(wrap_text=True)

    # ── The By Category sheet carries its own column set ────────────────────
    ri = ri + 2 + len(PURPOSE_GROUPS) + 1
    t = ws2.cell(row=ri, column=1, value="By Category sheet")
    t.font = Font(bold=True, size=11, color="3D6B2E")
    ws2.cell(row=ri, column=3,
             value="Long format: one row per run per dataset per subset, plus that run's "
                   "TOTAL on each dataset. Nothing is pooled, so Category is a label to "
                   "pivot on rather than a level the numbers were averaged into. Rows "
                   "whose Sub-dataset is TOTAL carry TOTAL in Category as well, so a "
                   "pivot excludes them with one filter. Metric columns are "
                   "the Failure Matrix's, unchanged, so a row carries values only in its own "
                   "dataset's columns and n/a in the other three. Only the three key "
                   "columns below are added; every metric column is documented above."
             ).alignment = Alignment(wrap_text=True)
    for ci, h in enumerate(["Stage", "Column", "Source / definition"], 1):
        c = ws2.cell(row=ri + 1, column=ci, value=h)
        c.font = Font(bold=True, size=10)
        c.fill = PatternFill("solid", fgColor="EDF3EA")
    for k, (grp, title, _key, desc) in enumerate(_LONG_KEY_COLS, ri + 2):
        ws2.cell(row=k, column=1, value=grp or "—")
        ws2.cell(row=k, column=2, value=title)
        ws2.cell(row=k, column=3, value=desc).alignment = Alignment(wrap_text=True)

    ws2.column_dimensions["A"].width = 18
    ws2.column_dimensions["B"].width = 24
    ws2.column_dimensions["C"].width = 80

    # ── Subset Map: which purpose group each official subset belongs to ─────
    ws4 = wb.create_sheet("Subset Map")
    heads = ["Dataset", "Sub-dataset", "Benchmark's own label", "Category",
             "Category name", "Notes"]
    for ci, h in enumerate(heads, 1):
        c = ws4.cell(row=1, column=ci, value=h)
        c.font = Font(bold=True)
        c.fill = PatternFill("solid", fgColor="F2F5F8")
    ri = 2
    for pre, raw, disp, code in _SUBSETS:
        ds = _LONG_DS_NAME[_DS_OF_PREFIX[pre]]
        ws4.cell(row=ri, column=1, value=ds)
        ws4.cell(row=ri, column=2, value=disp).font = Font(bold=True)
        ws4.cell(row=ri, column=3, value=raw if raw != disp else "")
        ws4.cell(row=ri, column=4, value=code).font = Font(bold=True)
        ws4.cell(row=ri, column=5, value=_GROUP_NAME[code])
        ws4.cell(row=ri, column=6,
                 value=_SUBSET_NOTE.get((pre, raw), "")).alignment = Alignment(wrap_text=True)
        if code in GROUP_FILL:
            for ci in range(1, 7):
                ws4.cell(row=ri, column=ci).fill = PatternFill(
                    "solid", fgColor=GROUP_FILL[code])
        ri += 1
    ws4.freeze_panes = "A2"
    ws4.auto_filter.ref = f"A1:F{ri - 1}"
    for col, w in zip("ABCDEF", (14, 26, 22, 10, 26, 78)):
        ws4.column_dimensions[col].width = w

    # ── Run parameters sheet ────────────────────────────────────────────────
    ws3 = wb.create_sheet("Run parameters")
    meta = [
        ["Generated at", datetime.now().strftime("%Y-%m-%d %H:%M")],
        ["Extraction LLM", "gemma-4-31B-it (openai-proxy/gemma-4-31B-it for Letta)"],
        ["Judge LLM", "gemma-4-E4B-it (OPENAI_MODEL)"],
        ["Judge temperature", "0 (pinned from this batch onward; earlier runs were not pinned)"],
        ["API endpoint", "NCHC GenAI portal"],
        ["", ""],
        ["LongMemEval scale", "18 questions (3 per type), haystack of about 50 sessions each"],
        ["LoCoMo scale", "1 conversation (conv-26), 19 sessions / 419 turns / 199 QA"],
        ["HaluMem scale", "1 user (the 2nd), 77 sessions / 188 QA"],
        ["MemFail scale", "35 questions (5 per subset, 15 for persona)"],
        ["", ""],
        ["top-k", "20 (identical across all four experiments)"],
        ["", ""],
        ["", "-- Cost and latency columns --"],
        ["Instrumentation", "halumem_experiment/token_tracker.py intercepts LLM calls in place while the experiment runs"],
        ["Buckets", "ingest (memory writes) / qa (retrieval + generation) / other (judge, probes, warm-up)"],
        ["Excluded", "Cost figures exclude the judge and the probes; those land in the other bucket and are not counted"],
        ["Definition of unit", "The span covered by one add call, i.e. the granularity column: turn or session"],
        ["Letta call counts", "Taken from usage.step_count; one messages.create may run several agent steps server-side"],
        ["Seconds per unit", "End-to-end median including retrieval and non-LLM overhead, not pure API time"],
        ["Why cells are blank", "Runs executed before 2026-08-18 recorded only total_tokens with no bucketing; they must be re-run to produce figures"],
        ["", ""],
        ["", "-- Run mapped to each backend --"],
    ]
    for name, hal, loc, lme, mf, llm in BACKENDS:
        meta.append([name, f"HaluMem={hal} | LoCoMo={loc} | LongMemEval={lme} | MemFail={mf}"])
    meta += [
        ["", ""],
        ["", "-- Known limitations --"],
        ["LongMemEval", "Not finished in this batch, so the column is blank"],
        ["HaluMem StructMem/A-MEM/Letta", "Reuses user2nd_gemma431b_probe from 2026-08-08 to 08-09, where judge temperature was not pinned"],
        ["LoCoMo Letta P1-scoped", "Letta has no turn provenance, so the scope falls back to the whole store and is not directly comparable to the other backends"],
        ["MemFail Letta", "results_5q_letta; the letta_b2 and letta_flush variants were not adopted"],
    ]
    for ri, (a, b) in enumerate(meta, 1):
        ws3.cell(row=ri, column=1, value=a).font = Font(bold=bool(a and not b.startswith("HaluMem=")))
        ws3.cell(row=ri, column=2, value=b).alignment = Alignment(wrap_text=True)
    ws3.column_dimensions["A"].width = 32
    ws3.column_dimensions["B"].width = 100

    wb.save(OUT)
    _write_outline_level_col(OUT)
    print(f"✅ {OUT}")
    filled = sum(1 for r in rows for k, v in r.items() if v is not None)
    print(f"   {len(rows)} rows x {len(COLUMNS)} columns, {filled} cells filled")
    return OUT


if __name__ == "__main__":
    build()
