"""
LongMemEval-S experiment_results.xlsx — APPEND/merge mode (keep history, never delete).
"""

import os
import re
import json
import glob
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE_DIR    = os.path.dirname(os.path.abspath(__file__))
RESULTS_DIR = os.path.join(BASE_DIR, "results")
EXCEL_PATH  = os.path.join(BASE_DIR, "experiment_results.xlsx")

# LongMemEval question types (short keys for columns)
QTYPES = ["multi-session", "temporal-reasoning", "knowledge-update",
          "single-session-user", "single-session-assistant", "single-session-preference"]
CFG="F2F3F4"; ACC="D6E4F0"; CAT="FEF9E7"; RET="E8DAEF"

COLUMNS = [
    ("Run", "run_name", 22, CFG),
    ("Date", "date", 11, CFG),
    ("Backend", "backend", 11, CFG),
    ("Granularity", "granularity", 11, CFG),
    ("Extraction LLM", "extraction_llm", 16, CFG),
    ("Judge LLM", "judge_llm", 15, CFG),
    ("Embedding", "embed_model", 15, CFG),
    ("Q#", "qa_num", 6, CFG),
    ("Scope", "scope", 13, CFG),
    ("Judge Acc", "acc_all", 10, ACC),
    ("Token F1", "token_f1", 9, ACC),
    ("multi-sess", "multi-session", 10, CAT),
    ("temporal", "temporal-reasoning", 10, CAT),
    ("know-update", "knowledge-update", 11, CAT),
    ("ss-user", "single-session-user", 9, CAT),
    ("ss-asst", "single-session-assistant", 9, CAT),
    ("ss-pref", "single-session-preference", 9, CAT),
    ("Recall@3", "recall@3", 9, RET),
    ("Recall@5", "recall@5", 9, RET),
    ("Recall@10", "recall@10", 9, RET),
    ("NDCG@5", "ndcg@5", 9, RET),
    ("P0 Trigger", "P0_trigger", 10, RET),
    ("KU-Rec@5", "ku_recall@5", 9, RET),
    ("KU-GoldCtx", "ku_gold_in_ctx", 10, RET),
    ("Tokens", "tokens", 11, CFG),
]


def _scope_label(qa_num):
    """Annotate a run's sample scale so full-30 vs small-sample rows aren't
    silently compared. Derived from question count."""
    if not qa_num:
        return "?"
    if qa_num >= 30:
        return "full-30"
    if qa_num == 6:
        return "1/type (6)"
    if qa_num <= 2:
        return "smoke"
    return f"partial ({qa_num})"


def _env_val(key, default="unknown"):
    env = os.path.join(BASE_DIR, "..", "halumem_experiment", ".env")
    if os.path.isfile(env):
        for line in open(env):
            m = re.match(rf"{key}\s*=\s*(.+)", line.strip())
            if m:
                return m.group(1).strip()
    return default


def _load_meta(run_dir, frame):
    p = os.path.join(run_dir, f"{frame}_lme_meta.json")
    if os.path.exists(p):
        try:
            return json.load(open(p))
        except Exception:
            pass
    return {"extraction_llm": _env_val("MEM0_LLM_MODEL"), "judge_llm": _env_val("OPENAI_MODEL"),
            "embed_model": _env_val("MEM0_EMBED_MODEL"), "granularity": "session"}


def scan_runs():
    runs = []
    for frame in ["mem0", "rag", "graphiti", "amem", "letta", "memos", "zep"]:
        for sp in sorted(glob.glob(os.path.join(RESULTS_DIR, f"{frame}-*", f"{frame}_lme_scores.json"))):
            run_dir = os.path.dirname(sp)
            version = os.path.basename(run_dir)[len(frame) + 1:]
            d = json.load(open(sp))
            tokp = os.path.join(run_dir, f"{frame}_lme_token_usage.json")
            tokens = json.load(open(tokp)).get("total_tokens") if os.path.exists(tokp) else None
            meta = _load_meta(run_dir, frame)
            row = {
                "run_name": f"{frame}-{version}",
                "date": datetime.fromtimestamp(os.path.getmtime(sp)).strftime("%Y-%m-%d"),
                "backend": {"mem0": "Mem0 OSS", "rag": "RAG", "graphiti": "Graphiti",
                            "amem": "A-MEM", "letta": "Letta", "memos": "MemOS",
                            "zep": "Zep Cloud"}.get(frame, frame),
                "granularity": meta.get("granularity"), "extraction_llm": meta.get("extraction_llm"),
                "judge_llm": meta.get("judge_llm"), "embed_model": meta.get("embed_model"),
                "qa_num": d.get("qa_num"), "acc_all": d.get("qa_accuracy_all"),
                "token_f1": d.get("token_f1_all"), "tokens": tokens,
                "scope": _scope_label(d.get("qa_num")),
            }
            for qt in QTYPES:
                row[qt] = d.get("per_type", {}).get(qt, {}).get("accuracy")
            for rk in ["recall@3", "recall@5", "recall@10", "ndcg@5"]:
                row[rk] = d.get("retrieval", {}).get(rk)
            for pk, pv in (d.get("probe_update") or {}).items():
                row[pk] = pv
            runs.append(row)
    return runs


def _read_existing():
    if not os.path.exists(EXCEL_PATH):
        return []
    try:
        ws = openpyxl.load_workbook(EXCEL_PATH).active
    except Exception:
        return []
    hdr = [c.value for c in ws[1]]
    h2k = {h: k for (h, k, _, _) in COLUMNS}
    out = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if not r or not r[0]:
            continue
        out.append({h2k[h]: r[i] for i, h in enumerate(hdr) if h in h2k and i < len(r)})
    return out


def build_excel(runs):
    merged = {r["run_name"]: r for r in _read_existing()}
    for r in runs:
        merged[r["run_name"]] = r
    ordered = list(merged.values())

    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "LongMemEval-S"
    thin = Side(style="thin", color="CCCCCC"); border = Border(thin, thin, thin, thin)
    for ci, (hdr, _, w, bg) in enumerate(COLUMNS, 1):
        c = ws.cell(row=1, column=ci, value=hdr)
        c.font = Font(bold=True, size=10); c.fill = PatternFill("solid", fgColor=bg)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = border
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[1].height = 28; ws.freeze_panes = "A2"
    for ri, run in enumerate(ordered, 2):
        for ci, (_, key, _, _) in enumerate(COLUMNS, 1):
            v = run.get(key); cell = ws.cell(row=ri, column=ci, value=v)
            cell.border = border; cell.alignment = Alignment(horizontal="center", vertical="center")
            if isinstance(v, float):
                cell.number_format = "0.000"
    wb.save(EXCEL_PATH)
    print(f"✅ LongMemEval Excel → {EXCEL_PATH}  ({len(ordered)} rows, {len(runs)} merged)")


if __name__ == "__main__":
    build_excel(scan_runs())
