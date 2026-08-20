"""
Build/refresh the LoCoMo experiment_results.xlsx (independent from halumem's).

Full refresh from each run's scores.json — idempotent. Columns:
  Run | Date | Backend | LLM | Conversations | QA# | Overall Acc
      | multi_hop | temporal | open_domain | single_hop | adversarial | Tokens
"""

import os
import re
import json
import glob
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE_DIR    = os.path.dirname(os.path.abspath(__file__))
RESULTS_DIR = os.path.join(BASE_DIR, "results")
LOGS_DIR    = os.path.join(BASE_DIR, "logs")
EXCEL_PATH  = os.path.join(BASE_DIR, "experiment_results.xlsx")

CATS = ["multi_hop", "temporal", "open_domain", "single_hop", "adversarial"]
HEADER_BG = "2E4057"; CFG_BG = "F2F3F4"; ACC_BG = "D6E4F0"; CAT_BG = "FEF9E7"

COLUMNS = [
    ("Run", "run_name", 24, CFG_BG),
    ("Date", "date", 11, CFG_BG),
    ("Backend", "backend", 11, CFG_BG),
    ("Granularity", "granularity", 12, CFG_BG),
    ("Extraction LLM", "extraction_llm", 18, CFG_BG),
    ("Judge LLM", "judge_llm", 16, CFG_BG),
    ("Embedding", "embed_model", 16, CFG_BG),
    ("Convs", "convs", 7, CFG_BG),
    ("QA#", "qa_num", 7, CFG_BG),
    # answer-quality
    ("Judge Acc", "acc_all", 11, ACC_BG),
    ("Token F1", "token_f1", 10, ACC_BG),
    # per-category judge accuracy
    ("multi_hop", "multi_hop", 11, CAT_BG),
    ("temporal", "temporal", 11, CAT_BG),
    ("open_domain", "open_domain", 12, CAT_BG),
    ("single_hop", "single_hop", 11, CAT_BG),
    ("adversarial", "adversarial", 12, CAT_BG),
    # retrieval
    # Extraction stage: a metric LoCoMo does not ship, computed with observations
    # as golden memories
    ("Integrity(R)", "extr_recall", 12, "D5F5E3"),
    ("Accuracy(P)", "extr_precision", 12, "D5F5E3"),
    ("Extraction F1", "extr_f1", 13, "D5F5E3"),
    ("SpkConfus%", "extr_spk_confusion", 11, "D5F5E3"),
    ("Extr Unk%", "extr_unknown", 10, "D5F5E3"),
    ("Extr Scope", "extr_scope", 11, "D5F5E3"),

    ("Recall@3", "recall@3", 10, "E8DAEF"),
    ("Recall@5", "recall@5", 10, "E8DAEF"),
    ("Recall@10", "recall@10", 10, "E8DAEF"),
    ("NDCG@3", "ndcg@3", 10, "E8DAEF"),
    ("NDCG@5", "ndcg@5", 10, "E8DAEF"),
    ("NDCG@10", "ndcg@10", 10, "E8DAEF"),
    ("Tokens", "tokens", 11, CFG_BG),
]


def _env_val(key, default="unknown"):
    """Read a value from halumem's .env (shared config)."""
    env = os.path.join(BASE_DIR, "..", "halumem_experiment", ".env")
    if os.path.isfile(env):
        for line in open(env):
            m = re.match(rf"{key}\s*=\s*(.+)", line.strip())
            if m:
                return m.group(1).strip()
    return default


def _load_meta(run_dir, frame):
    """Load per-run config meta; fall back to .env for older runs."""
    path = os.path.join(run_dir, f"{frame}_locomo_meta.json")
    if os.path.exists(path):
        try:
            return json.load(open(path))
        except Exception:
            pass
    return {
        "extraction_llm": _env_val("MEM0_LLM_MODEL"),
        "judge_llm":      _env_val("OPENAI_MODEL"),
        "embed_model":    _env_val("MEM0_EMBED_MODEL"),
        "granularity":    "session",
    }


def scan_runs():
    runs = []
    for frame in ["mem0", "rag", "graphiti", "amem", "letta"]:
        for scores_path in sorted(glob.glob(os.path.join(RESULTS_DIR, f"{frame}-*", f"{frame}_locomo_scores.json"))):
            run_dir = os.path.dirname(scores_path)
            version = os.path.basename(run_dir)[len(frame) + 1:]
            with open(scores_path, encoding="utf-8") as f:
                d = json.load(f)
            tok_path = os.path.join(run_dir, f"{frame}_locomo_token_usage.json")
            tokens = None
            if os.path.exists(tok_path):
                try:
                    tokens = json.load(open(tok_path)).get("total_tokens")
                except Exception:
                    pass
            tmp = os.path.join(run_dir, "tmp")
            convs = len([f for f in os.listdir(tmp) if f.endswith(".json")]) if os.path.isdir(tmp) else 0
            meta = _load_meta(run_dir, frame)
            row = {
                "run_name": f"{frame}-{version}",
                "date": datetime.fromtimestamp(os.path.getmtime(scores_path)).strftime("%Y-%m-%d"),
                "backend": {"mem0": "Mem0 OSS", "rag": "RAG", "graphiti": "Graphiti", "amem": "A-MEM", "letta": "Letta"}.get(frame, frame),
                "granularity":    meta.get("granularity"),
                "extraction_llm": meta.get("extraction_llm"),
                "judge_llm":      meta.get("judge_llm"),
                "embed_model":    meta.get("embed_model"),
                "convs": convs,
                "qa_num": d.get("qa_num"),
                "acc_all": d.get("qa_accuracy_all"),
                "token_f1": d.get("token_f1_all"),
                "tokens": tokens,
            }
            # Extraction stage. A run without a memory_dump yields {"skipped": ...},
            # so leave every cell blank rather than writing 0, which would be read as
            # "extraction failed completely".
            ex = d.get("extraction") or {}
            row["extr_recall"]         = ex.get("memory_integrity_recall")
            row["extr_precision"]      = ex.get("memory_accuracy_precision")
            row["extr_f1"]             = ex.get("memory_extraction_f1")
            row["extr_spk_confusion"]  = ex.get("speaker_confusion_ratio")
            row["extr_unknown"]        = ex.get("unknown_ratio")
            scope = ex.get("scope")
            row["extr_scope"] = ("／".join(scope) if isinstance(scope, list) else scope)

            for c in CATS:
                row[c] = d.get("per_category", {}).get(c, {}).get("accuracy")
            for rk in ["recall@3", "recall@5", "recall@10", "ndcg@3", "ndcg@5", "ndcg@10"]:
                row[rk] = d.get("retrieval", {}).get(rk)
            runs.append(row)
    return runs


def _read_existing():
    """Read existing Excel rows into dicts (by column key). Preserves history even
    if a run's results folder was later deleted."""
    if not os.path.exists(EXCEL_PATH):
        return []
    try:
        ws = openpyxl.load_workbook(EXCEL_PATH).active
    except Exception:
        return []
    hdr = [c.value for c in ws[1]]
    hdr_to_key = {h: k for (h, k, _, _) in COLUMNS}
    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if not r or not r[0]:
            continue
        row = {}
        for i, h in enumerate(hdr):
            key = hdr_to_key.get(h)
            if key and i < len(r):
                row[key] = r[i]
        rows.append(row)
    return rows


def build_excel(runs):
    """APPEND / merge mode: keep every existing row, update-or-add the given runs
    by run_name, never delete. So each run accumulates into the file permanently."""
    merged = {r["run_name"]: r for r in _read_existing()}
    for r in runs:                      # update existing or append new
        merged[r["run_name"]] = r
    ordered = list(merged.values())

    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "LoCoMo"
    thin = Side(style="thin", color="CCCCCC"); border = Border(thin, thin, thin, thin)
    for ci, (hdr, _, w, bg) in enumerate(COLUMNS, 1):
        c = ws.cell(row=1, column=ci, value=hdr)
        c.font = Font(bold=True, color="FFFFFF" if bg == HEADER_BG else "000000", size=10)
        c.fill = PatternFill("solid", fgColor=bg)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = border
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[1].height = 26
    ws.freeze_panes = "A2"

    for ri, run in enumerate(ordered, 2):
        for ci, (_, key, _, _) in enumerate(COLUMNS, 1):
            v = run.get(key)
            cell = ws.cell(row=ri, column=ci, value=v)
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if isinstance(v, float):
                cell.number_format = "0.000"
    wb.save(EXCEL_PATH)
    print(f"✅ LoCoMo Excel → {EXCEL_PATH}  ({len(ordered)} total rows, {len(runs)} run(s) merged)")


if __name__ == "__main__":
    build_excel(scan_runs())
