"""
MemFail experiment_results.xlsx -- LONG format.

Layout (fixed, all future runs follow it):
    column 1 = Run Name
    column 2 = Dataset
    -> every run occupies exactly 6 rows: the 5 MemFail sub-datasets plus a
       TOTAL row. Sub-datasets with no data still get a row (blank metrics) so
       the block height is constant and runs stay comparable side by side.

Scores come from each sub-dataset's analysis_*.csv (the official
analyze_errors.py output); run settings come from graded_traces run_metadata.

MemFail's attribution SHORT-CIRCUITS on the first failing stage, so a blank
later stage means "not checked", not "passed". Only Acc and the first failing
stage are safe to read directly.

Merge policy: append/merge keyed on (Run Name, Dataset) -- rerunning one run
never drops the others.
"""

import os
import csv
import json
import glob
import subprocess
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
EXCEL_PATH = os.path.join(BASE_DIR, "experiment_results.xlsx")
MD_PATH = os.path.join(BASE_DIR, "experiment_results.md")

# results dir key -> display name (order fixed; every run shows all of them)
DATASETS = [
    ("coexisting_facts", "coexisting_facts"),
    ("conditional_easy", "conditional_easy"),
    ("conditional_hard", "conditional_hard"),
    ("long_hop",         "long_hop"),
    ("persona",          "persona_retrieval"),
]
TOTAL_LABEL = "TOTAL"

COLUMNS = [
    ("Run Name",     "run_name",     24, "cfg"),
    ("Dataset",      "dataset",      20, "cfg"),
    ("Date",         "date",         11, "cfg"),
    ("Memory",       "memory",       10, "cfg"),
    ("Mem0 Ver",     "mem0_version", 10, "cfg"),
    ("Extract LLM",  "extract_llm",  17, "cfg"),
    ("Answer LLM",   "answer_llm",   16, "cfg"),
    ("Judge LLM",    "judge_llm",    16, "cfg"),
    ("Embedding",    "embedding",    15, "cfg"),
    ("Q#",           "n",             6, "score"),
    ("Correct",      "correct",       8, "score"),
    ("Acc",          "acc",           8, "score"),
    ("Storage Err",  "storage",      11, "err"),
    ("Summary Err",  "summary",      11, "err"),
    ("Retr Err",     "retrieval",    10, "err"),
    ("Reason Err",   "reasoning",    11, "err"),
    ("Store Size",   "store_size",   10, "score"),
]

C_HEADER_BG = "2E4057"; C_HEADER_FG = "FFFFFF"
C_CFG   = "F2F3F4"
C_SCORE = "D6E4F0"
C_ERR   = "FEF9E7"
C_RUN_A = "FFFFFF"; C_RUN_B = "F8F9FA"
C_TOTAL = "E1EFED"
BG = {"cfg": C_CFG, "score": C_SCORE, "err": C_ERR}


def _latest(pattern):
    fs = sorted(glob.glob(pattern))
    return fs[-1] if fs else None


def _probe_mem0_version(venv):
    try:
        out = subprocess.run(
            [f"../{venv}/bin/python", "-c", "import mem0;print(mem0.__version__)"],
            capture_output=True, text=True, cwd=BASE_DIR, timeout=30
        ).stdout.strip()
        return out or None
    except Exception:
        return None


def _mem0_version(root, run_name):
    """v1 and v2 runs come from different venvs, so probing the current
    interpreter would stamp every run with whichever one is installed here.
    The run itself records its version; name and probe are only fallbacks."""
    marker = os.path.join(root, "mem0_version.txt")
    if os.path.isfile(marker):
        v = open(marker).read().strip()
        if v:
            return v
    if "mem0v1" in run_name:
        return _probe_mem0_version("venv_mem0v1")
    if "mem0v2" in run_name:
        return _probe_mem0_version("venv_memfail")
    return _probe_mem0_version("venv_memfail")


def scan_runs():
    """Return long-format rows: one per (run, dataset) plus a TOTAL per run."""
    rows = []

    for root in sorted(glob.glob(os.path.join(BASE_DIR, "results_*"))):
        run_name = os.path.basename(root)
        cfg = {"run_name": run_name, "mem0_version": _mem0_version(root, run_name)}
        per_ds = []

        for key, label in DATASETS:
            rec = {"dataset": label, "n": None, "correct": None, "acc": None,
                   "storage": None, "summary": None, "retrieval": None,
                   "reasoning": None, "store_size": None}

            gt = _latest(os.path.join(root, key, "*", "graded_traces_*.json"))
            if gt:
                d = json.load(open(gt))
                md = d.get("run_metadata", {}) or {}
                cli = md.get("all_cli_args", {}) or {}
                mem = md.get("memory_system")
                # Every backend's CLI flags are always present with their
                # defaults, so read only the ones the selected backend uses --
                # otherwise an A-MEM run reports mem0's default model.
                per_backend = {
                    "mem0":      ("mem0_llm_model", "mem0_embedding_model"),
                    "amem":      ("amem_llm_model", "amem_embedding_model"),
                    "structmem": ("structmem_model", "structmem_embedding_model"),
                    "simplemem": ("simplemem_model", None),
                    # Letta's writer is the agent itself, so its agent LLM is
                    # the extraction model; its embedding is a server-side handle.
                    "letta":     ("letta_llm_model", "letta_embedding_model"),
                }
                lk, ek = per_backend.get(mem, ("mem0_llm_model", "mem0_embedding_model"))
                cfg.setdefault("date", datetime.fromtimestamp(os.path.getmtime(gt)).strftime("%Y-%m-%d"))
                cfg.setdefault("memory", mem)
                cfg.setdefault("extract_llm", cli.get(lk))
                cfg.setdefault("answer_llm", md.get("llm_model"))
                # analyze_errors.py does not record its judge model in the CSV;
                # every sweep here used gemma-4-31B-it as judge.
                cfg.setdefault("judge_llm", os.getenv("MEMFAIL_JUDGE", "gemma-4-31B-it"))
                cfg.setdefault("embedding", cli.get(ek) if ek else None)
                # mem0 version only means anything for the mem0 backend
                if mem != "mem0":
                    cfg["mem0_version"] = None
                rec["store_size"] = len(d.get("all_memories_at_time_of_questions") or [])

            an = _latest(os.path.join(root, key, "analysis", "analysis_*.csv"))
            if an:
                ar = list(csv.DictReader(open(an)))
                n = len(ar)
                if n:
                    correct = sum(1 for r in ar if (r.get("judge_result") or "").strip() == "correct")
                    # analyze_errors.py labels the storage and retrieval stages
                    # "not_stored"/"not_retrieved"; only summary and reasoning
                    # use the *_error suffix. Matching on the suffix alone left
                    # those two columns reading zero on every run.
                    stage_of = {
                        "not_stored": "storage", "storage_error": "storage",
                        "summary_error": "summary",
                        "not_retrieved": "retrieval", "retrieval_error": "retrieval",
                        "reasoning_error": "reasoning",
                    }
                    counts = {"storage": 0, "summary": 0, "retrieval": 0, "reasoning": 0}
                    for r in ar:
                        if (r.get("judge_result") or "").strip() == "correct":
                            continue
                        st = stage_of.get((r.get("error_type") or "").strip())
                        if st:
                            counts[st] += 1
                    rec.update({"n": n, "correct": correct,
                                "acc": round(correct / n, 4), **counts})
            per_ds.append(rec)

        if not any(r["n"] for r in per_ds):
            continue

        tot_n = sum(r["n"] or 0 for r in per_ds)
        tot_c = sum(r["correct"] or 0 for r in per_ds)
        total = {
            "dataset": TOTAL_LABEL, "n": tot_n, "correct": tot_c,
            "acc": round(tot_c / tot_n, 4) if tot_n else None,
            "storage":   sum(r["storage"] or 0 for r in per_ds),
            "summary":   sum(r["summary"] or 0 for r in per_ds),
            "retrieval": sum(r["retrieval"] or 0 for r in per_ds),
            "reasoning": sum(r["reasoning"] or 0 for r in per_ds),
            "store_size": sum(r["store_size"] or 0 for r in per_ds),
        }
        for rec in per_ds + [total]:
            rows.append({**cfg, **rec})
    return rows


def _read_existing():
    if not os.path.exists(EXCEL_PATH):
        return {}
    try:
        ws = openpyxl.load_workbook(EXCEL_PATH).active
    except Exception:
        return {}
    hdr = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    h2k = {h: k for h, k, _, _ in COLUMNS}
    out = {}
    for ri in range(2, ws.max_row + 1):
        rec = {h2k[h]: ws.cell(ri, ci).value
               for ci, h in enumerate(hdr, 1) if h in h2k}
        if rec.get("run_name") and rec.get("dataset"):
            out[(rec["run_name"], rec["dataset"])] = rec
    return out


def build_excel(rows):
    merged = _read_existing()
    for r in rows:
        merged[(r["run_name"], r["dataset"])] = r

    # keep runs grouped, datasets in canonical order, TOTAL last
    order = {label: i for i, (_k, label) in enumerate(DATASETS)}
    order[TOTAL_LABEL] = len(order)
    ordered = sorted(merged.values(),
                     key=lambda r: (str(r.get("run_name")), order.get(r.get("dataset"), 99)))

    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "MemFail"
    thin = Side(style="thin", color="CCCCCC"); border = Border(thin, thin, thin, thin)

    for ci, (h, _k, w, kind) in enumerate(COLUMNS, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font = Font(bold=True, size=9, color=C_HEADER_FG)
        c.fill = PatternFill("solid", fgColor=C_HEADER_BG)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = border
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[1].height = 26
    ws.freeze_panes = "C2"

    run_ix = {}
    for ri, r in enumerate(ordered, start=2):
        rn = str(r.get("run_name"))
        run_ix.setdefault(rn, len(run_ix))
        is_total = r.get("dataset") == TOTAL_LABEL
        base = C_TOTAL if is_total else (C_RUN_A if run_ix[rn] % 2 == 0 else C_RUN_B)
        for ci, (_h, k, _w, kind) in enumerate(COLUMNS, 1):
            v = r.get(k)
            cell = ws.cell(row=ri, column=ci, value=v)
            cell.fill = PatternFill("solid", fgColor=base)
            cell.border = border
            cell.alignment = Alignment(
                horizontal="left" if k in ("run_name", "dataset") else "center",
                vertical="center")
            if is_total:
                cell.font = Font(bold=True, size=9)
            if isinstance(v, float):
                cell.number_format = "0.000"
        ws.row_dimensions[ri].height = 16

    wb.save(EXCEL_PATH)
    print(f"MemFail Excel -> {EXCEL_PATH}  ({len(ordered)} rows, {len(run_ix)} runs)")


def build_markdown(rows):
    lines = ["# MemFail experiment results", "",
             "| Run | Dataset | Q# | Correct | Acc | Storage | Summary | Retr | Reason | Store |",
             "|---|---|---|---|---|---|---|---|---|---|"]
    for r in rows:
        f = lambda k: "" if r.get(k) is None else r.get(k)
        lines.append(
            f"| {r['run_name']} | {r['dataset']} | {f('n')} | {f('correct')} | {f('acc')} | "
            f"{f('storage')} | {f('summary')} | {f('retrieval')} | {f('reasoning')} | {f('store_size')} |"
        )
    lines += ["", "## Run settings", ""]
    seen = set()
    for r in rows:
        if r["run_name"] in seen:
            continue
        seen.add(r["run_name"])
        lines.append(f"- **{r['run_name']}** ({r.get('date')}) — memory `{r.get('memory')}` "
                     f"(mem0 {r.get('mem0_version')}), extract `{r.get('extract_llm')}`, "
                     f"answer `{r.get('answer_llm')}`, judge `{r.get('judge_llm')}`, "
                     f"embed `{r.get('embedding')}`")
    lines += ["", "> Attribution short-circuits on the first failing stage; a blank later stage "
                  "means 'not checked', not 'passed'."]
    open(MD_PATH, "w").write("\n".join(lines))
    print(f"Markdown -> {MD_PATH}")


if __name__ == "__main__":
    rs = scan_runs()
    build_excel(rs)
    build_markdown(rs)
