"""
update_excel.py: scan every results/ directory and create or update experiment_results.xlsx.

Usage:
  python update_excel.py              # update manually
  python update_excel.py --open       # update, then open in Excel

Also called automatically at the end of every run.py invocation.
"""

import os
import re
import json
import glob
import argparse
import subprocess
from datetime import datetime

import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

BASE_DIR    = os.path.dirname(os.path.abspath(__file__))
RESULTS_DIR = os.path.join(BASE_DIR, "results")
LOGS_DIR    = os.path.join(BASE_DIR, "logs")
EXCEL_PATH  = os.path.join(BASE_DIR, "experiment_results.xlsx")

# ── Colour palette ──────────────────────────────────────────────────────────
C_HEADER_BG   = "2E4057"   # dark blue-grey
C_HEADER_FG   = "FFFFFF"
C_SEC1_BG     = "D6E4F0"   # light blue  — Memory Point
C_SEC2_BG     = "D5F5E3"   # light green — Memory Update
C_SEC3_BG     = "FEF9E7"   # light yellow— QA
C_SEC4_BG     = "FADBD8"   # light red    — QA Attribution
C_CONFIG_BG   = "F2F3F4"   # grey         — config columns
C_ROW_ODD     = "FFFFFF"
C_ROW_EVEN    = "F8F9FA"
C_BEST_FG     = "1A5276"   # dark blue for best value highlight


def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)


def _font(bold=False, color="000000", size=10):
    return Font(bold=bold, color=color, size=size)


def _border():
    thin = Side(style="thin", color="CCCCCC")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def _pct(value):
    """Format float as percentage string for display."""
    if value is None:
        return None
    return round(float(value), 4)


# ── Config extraction ────────────────────────────────────────────────────────

def _extract_config_from_log(log_dir: str) -> dict:
    """Read first line of any .log file in log_dir to get LLM names."""
    config = {"extraction_llm": "unknown", "backend_detail": ""}
    if not os.path.isdir(log_dir):
        return config
    logs = [f for f in os.listdir(log_dir) if f.endswith(".log")]
    if not logs:
        return config
    first_line = ""
    with open(os.path.join(log_dir, logs[0]), encoding="utf-8") as f:
        first_line = f.readline().strip()

    # RAG baseline has no extraction LLM by design (stores raw turns).
    if re.search(r"\|\s*RAG\s*\|", first_line):
        config["extraction_llm"] = "none (raw turns)"
        return config

    # Mem0: "=== Start user {name} ({uuid}) | MEM0_LLM={model} ==="
    m = re.search(r"MEM0_LLM=([^\s|=]+)", first_line)
    if m:
        config["extraction_llm"] = m.group(1)
        return config

    # Graphiti: "=== Start {name} ({uuid}) | LLM={model} ==="
    m = re.search(r"LLM=([^\s|=]+)", first_line)
    if m:
        config["extraction_llm"] = m.group(1)
    return config


def _extract_run_meta(run_dir: str, frame: str) -> dict:
    """Get user count, total sessions, run date from tmp files."""
    tmp_dir = os.path.join(run_dir, "tmp")
    meta = {"users": 0, "sessions": 0, "run_date": ""}
    if not os.path.isdir(tmp_dir):
        return meta

    jsons = [f for f in os.listdir(tmp_dir) if f.endswith(".json")]
    meta["users"] = len(jsons)

    total_sessions = 0
    earliest_mtime = None
    for fname in jsons:
        fpath = os.path.join(tmp_dir, fname)
        mtime = os.path.getmtime(fpath)
        if earliest_mtime is None or mtime < earliest_mtime:
            earliest_mtime = mtime
        try:
            with open(fpath, encoding="utf-8") as f:
                d = json.load(f)
            sessions = d.get("sessions", [])
            total_sessions += sum(
                1 for s in sessions
                if not s.get("is_generated_qa_session", False)
            )
        except Exception:
            pass

    meta["sessions"] = total_sessions
    if earliest_mtime:
        meta["run_date"] = datetime.fromtimestamp(earliest_mtime).strftime("%Y-%m-%d")
    return meta


def _infer_embed_model(run_dir: str, frame: str) -> str:
    """Infer embedding model from env or directory name."""
    env_path = os.path.join(BASE_DIR, ".env")
    if os.path.isfile(env_path):
        with open(env_path) as f:
            content = f.read()
        if frame == "mem0_oss":
            m = re.search(r"MEM0_EMBED_MODEL\s*=\s*(.+)", content)
        elif frame == "memwave":
            m = re.search(r"MEMWEAVE_EMBED_MODEL\s*=\s*(.+)", content)
        elif frame == "amem":
            m = re.search(r"AMEM_EMBED_MODEL\s*=\s*(.+)", content)
        else:
            m = re.search(r"GRAPHITI_EMBED_MODEL\s*=\s*(.+)", content)
        if m:
            return m.group(1).strip()
    return "mxbai-embed-large"


def _infer_judge_llm() -> str:
    """Read current OPENAI_MODEL from .env as judge LLM."""
    env_path = os.path.join(BASE_DIR, ".env")
    if os.path.isfile(env_path):
        with open(env_path) as f:
            for line in f:
                m = re.match(r"OPENAI_MODEL\s*=\s*(.+)", line.strip())
                if m:
                    return m.group(1).strip()
    return "unknown"


# ── Score loading ─────────────────────────────────────────────────────────────

def load_scores(scores_path: str) -> dict:
    with open(scores_path, encoding="utf-8") as f:
        d = json.load(f)

    # Load token usage if available (recorded since 2026-05-26)
    token_path = scores_path.replace("_scores.json", "_token_usage.json")
    total_tokens = None
    if os.path.exists(token_path):
        try:
            with open(token_path, encoding="utf-8") as f:
                total_tokens = json.load(f).get("total_tokens")
        except Exception:
            pass

    attr = d.get("qa_attribution", {})

    return {
        "recall":                d["memory_integrity"]["recall(all)"],
        "precision":             d["memory_accuracy"]["target_accuracy(all)"],
        "f1":                    d["memory_extraction_f1"],
        "interference_accuracy": d["memory_accuracy"]["interference_accuracy(all)"],
        "weighted_accuracy":     d["memory_accuracy"]["weighted_accuracy(all)"],
        "update_correct":        d["memory_update"]["correct_update_memory_ratio(all)"],
        "update_omission":       d["memory_update"]["omission_update_memory_ratio(all)"],
        "update_halluc":         d.get("memory_update", {}).get("hallucination_update_memory_ratio(all)"),
        **{k: v for k, v in (d.get("probe_update") or {}).items() if isinstance(v, (int, float))},
        "update_halluc":         d.get("memory_update", {}).get("hallucination_update_memory_ratio(all)"),
        **{k: v for k, v in (d.get("probe_update") or {}).items() if isinstance(v, (int, float))},
        "qa_correct":            d["question_answering"]["correct_qa_ratio(all)"],
        "qa_hallucination":      d["question_answering"]["hallucination_qa_ratio(all)"],
        "qa_omission":           d["question_answering"]["omission_qa_ratio(all)"],
        "total_tokens":          total_tokens,
        # QA error attribution (None if this run was scored before attribution existed)
        "attr_wrong":            attr.get("total_wrong"),
        "attr_storage":          attr.get("storage_ratio"),
        "attr_retrieval":        attr.get("retrieval_ratio"),
        "attr_generation":       attr.get("generation_ratio"),
        # Unadjudicable share: the three ratios above already exclude these, and
        # without it the reliability of the attribution cannot be assessed
        "attr_unknown":          attr.get("unknown_ratio"),
    }


# ── Scan all runs ─────────────────────────────────────────────────────────────

def scan_runs() -> list[dict]:
    runs = []

    for frame in ["mem0_oss", "graphiti", "amem", "memwave", "rag", "letta", "zep", "memos", "structmem"]:
        pattern = os.path.join(RESULTS_DIR, f"{frame}-*", f"{frame}_scores.json")
        for scores_path in sorted(glob.glob(pattern)):
            run_dir = os.path.dirname(scores_path)
            version = os.path.basename(run_dir)[len(frame) + 1:]  # strip "frame-"
            log_dir = os.path.join(LOGS_DIR, f"{frame}-{version}")

            config  = _extract_config_from_log(log_dir)
            meta    = _extract_run_meta(run_dir, frame)
            scores  = load_scores(scores_path)

            runs.append({
                "run_name":       f"{frame}-{version}",
                "run_date":       meta["run_date"],
                "backend":        {"mem0_oss": "Mem0 OSS", "graphiti": "Graphiti", "amem": "A-MEM", "memwave": "MemWave", "rag": "RAG", "letta": "Letta", "zep": "Zep Cloud", "memos": "MemOS", "structmem": "StructMem"}.get(frame, frame),
                "extraction_llm": config["extraction_llm"],
                "embed_model":    _infer_embed_model(run_dir, frame),
                "judge_llm":      _infer_judge_llm(),
                "dataset":        "HaluMem-Medium",
                "users":          meta["users"],
                "sessions":       meta["sessions"],
                **scores,
            })

    return runs


# ── Excel builder ─────────────────────────────────────────────────────────────

COLUMNS = [
    # (header, field_key, width, section_color, is_pct)
    ("Run Name",        "run_name",         28, C_CONFIG_BG, False),
    ("Date",            "run_date",         12, C_CONFIG_BG, False),
    ("Backend",         "backend",          14, C_CONFIG_BG, False),
    ("Extraction LLM",  "extraction_llm",   28, C_CONFIG_BG, False),
    ("Embed Model",     "embed_model",      20, C_CONFIG_BG, False),
    ("Judge LLM",       "judge_llm",        22, C_CONFIG_BG, False),
    ("Dataset",         "dataset",          16, C_CONFIG_BG, False),
    ("Users",           "users",             8, C_CONFIG_BG, False),
    ("Sessions",        "sessions",         10, C_CONFIG_BG, False),
    ("Total Tokens",    "total_tokens",     14, C_CONFIG_BG, False),
    # Memory Point
    ("Recall",            "recall",                10, C_SEC1_BG,   True),
    ("Precision",         "precision",             10, C_SEC1_BG,   True),
    ("F1",                "f1",                    10, C_SEC1_BG,   True),
    ("Interference Acc",  "interference_accuracy", 16, C_SEC1_BG,   True),
    ("Weighted Acc",      "weighted_accuracy",     12, C_SEC1_BG,   True),
    # Memory Update
    ("Correct",           "update_correct",        10, C_SEC2_BG,   True),
    ("Omission",          "update_omission",       10, C_SEC2_BG,   True),
    ("Halluc",            "update_halluc",         10, C_SEC2_BG,   True),
    # QA
    ("Correct",           "qa_correct",            10, C_SEC3_BG,   True),
    ("Hallucination",     "qa_hallucination",      14, C_SEC3_BG,   True),
    ("Omission",          "qa_omission",           10, C_SEC3_BG,   True),
    # QA Error Attribution
    ("Wrong #",           "attr_wrong",            10, C_SEC4_BG,   False),
    ("Storage %",         "attr_storage",          11, C_SEC4_BG,   True),
    ("Retrieval %",       "attr_retrieval",        12, C_SEC4_BG,   True),
    ("Generation %",      "attr_generation",       13, C_SEC4_BG,   True),
    ("Unknown %",         "attr_unknown",          11, C_SEC4_BG,   True),
    # Update Probe (P0-P4)
    ("P0 Trigger",        "P0_trigger",            11, C_SEC1_BG,   True),
    ("P1 Loose",          "P1_loose",              10, C_SEC1_BG,   True),
    ("P1 Strict",         "P1_strict",             10, C_SEC1_BG,   True),
    ("P1 Distort",        "P1_distort",            11, C_SEC1_BG,   True),
    ("P1u OK",            "P1u_ok",                10, C_SEC2_BG,   True),
    ("P4 NewRet",         "P4u_new_retrieved",     11, C_SEC3_BG,   True),
    ("P4b StaleCtx",      "P4b_stale_in_ctx",      13, C_SEC3_BG,   True),
]

# Section header spans
SECTIONS = [
    ("Configuration",     1, 10),
    ("Memory Point",     11, 15),
    ("Memory Update",    16, 18),
    ("QA",               19, 21),
    ("QA Attribution",   22, 25),
    ("Update Probe",     26, 32),
]


def _write_row(ws, row_idx: int, run: dict):
    """Write a single data row with formatting."""
    bg = C_ROW_ODD if (row_idx % 2 == 1) else C_ROW_EVEN
    for col_idx, (_, field, _, _, is_pct) in enumerate(COLUMNS, start=1):
        value = run.get(field)
        cell  = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.fill      = _fill(bg)
        cell.border    = _border()
        cell.alignment = Alignment(
            horizontal="center" if is_pct else "left",
            vertical="center"
        )
        if is_pct and value is not None:
            cell.number_format = "0.000"
        if field in ("run_name", "extraction_llm", "embed_model", "judge_llm"):
            cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row_idx].height = 18


def _existing_run_names(ws) -> list[str]:
    """Read run names already in the sheet (column 1, rows 3+)."""
    names = []
    for row in ws.iter_rows(min_row=3, max_col=1, values_only=True):
        val = row[0]
        if val:
            names.append(str(val))
    return names


def _init_headers(ws):
    """Write section + column headers into a fresh sheet."""
    for title, col_start, col_end in SECTIONS:
        cell = ws.cell(row=1, column=col_start, value=title)
        cell.font      = _font(bold=True, color=C_HEADER_FG, size=11)
        cell.fill      = _fill(C_HEADER_BG)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = _border()
        if col_start != col_end:
            ws.merge_cells(
                start_row=1, start_column=col_start,
                end_row=1,   end_column=col_end
            )
    for col_idx, (header, _, width, sec_color, _) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        cell.font      = _font(bold=True, color="000000", size=10)
        cell.fill      = _fill(sec_color)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = _border()
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 30
    ws.freeze_panes = "A3"


def build_excel(runs: list[dict]):
    """Full refresh: rebuild the whole sheet from scratch every time.

    Because every value is derived from each run's scores.json / logs, a full
    rebuild is lossless and idempotent. This makes new runs, updated values
    (e.g. backfilled QA attribution, re-evals), and new columns/sections all
    show up automatically — no manual delete-and-rebuild needed.
    """
    # Preserve a stable, readable order: existing row order first (if any),
    # then any brand-new runs appended after.
    prior_order = []
    if os.path.exists(EXCEL_PATH):
        try:
            prior_order = _existing_run_names(openpyxl.load_workbook(EXCEL_PATH).active)
        except Exception:
            prior_order = []

    by_name = {r["run_name"]: r for r in runs}
    ordered_names = [n for n in prior_order if n in by_name]
    ordered_names += [r["run_name"] for r in runs if r["run_name"] not in ordered_names]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Results"
    _init_headers(ws)

    row_idx = 3
    for name in ordered_names:
        _write_row(ws, row_idx, by_name[name])
        row_idx += 1

    wb.save(EXCEL_PATH)
    new_count = len([n for n in ordered_names if n not in prior_order])
    print(f"✅ Excel refreshed → {EXCEL_PATH}  ({len(ordered_names)} runs, "
          f"{new_count} new, all values updated)")


# ── Markdown builder ──────────────────────────────────────────────────────────

MD_PATH = os.path.join(BASE_DIR, "experiment_results.md")

HIGHER_IS_BETTER = {"recall", "precision", "f1", "interference_accuracy", "weighted_accuracy", "update_correct", "qa_correct"}
LOWER_IS_BETTER  = {"update_omission", "qa_hallucination", "qa_omission"}


def _best_values(runs: list[dict]) -> dict:
    """Return {field: best_value} for all metric fields."""
    metric_fields = [f for _, f, _, _, is_pct in COLUMNS if is_pct]
    bests = {}
    for field in metric_fields:
        vals = [r[field] for r in runs if r.get(field) is not None]
        if not vals:
            continue
        if field in HIGHER_IS_BETTER:
            bests[field] = max(vals)
        elif field in LOWER_IS_BETTER:
            bests[field] = min(vals)
    return bests


def build_markdown(runs: list[dict]):
    bests = _best_values(runs)
    lines = []
    lines.append(f"# HaluMem Experiment Results")
    lines.append(f"")
    lines.append(f"_Last updated: {datetime.now().strftime('%Y-%m-%d %H:%M')}_")
    lines.append(f"")

    # ── Config table ──────────────────────────────────────────────────────
    lines.append("## Configuration")
    lines.append("")
    lines.append("| Run | Date | Backend | Extraction LLM | Embed Model | Judge LLM | Users | Sessions |")
    lines.append("|-----|------|---------|----------------|-------------|-----------|------:|--------:|")
    for r in runs:
        lines.append(
            f"| `{r['run_name']}` "
            f"| {r['run_date']} "
            f"| {r['backend']} "
            f"| `{r['extraction_llm']}` "
            f"| `{r['embed_model']}` "
            f"| `{r['judge_llm']}` "
            f"| {r['users']} "
            f"| {r['sessions']} |"
        )

    lines.append("")

    # ── Results table ─────────────────────────────────────────────────────
    lines.append("## Results")
    lines.append("")
    lines.append("| Run | Total Tokens | MP Recall | MP Precision | MP F1 | Interference Acc | Weighted Acc | Upd Correct | Upd Omission | QA Correct | QA Halluc | QA Omission |")
    lines.append("|-----|-------------:|----------:|-------------:|------:|-----------------:|-------------:|------------:|-------------:|-----------:|----------:|------------:|")

    def fmt(field, value):
        if value is None:
            return "—"
        s = f"{value:.3f}"
        if bests.get(field) == value:
            return f"**{s}**"   # bold = best in column
        return s

    for r in runs:
        tok = f"{r['total_tokens']:,}" if r.get('total_tokens') is not None else "—"
        lines.append(
            f"| `{r['run_name']}` "
            f"| {tok} "
            f"| {fmt('recall', r['recall'])} "
            f"| {fmt('precision', r['precision'])} "
            f"| {fmt('f1', r['f1'])} "
            f"| {fmt('interference_accuracy', r['interference_accuracy'])} "
            f"| {fmt('weighted_accuracy', r['weighted_accuracy'])} "
            f"| {fmt('update_correct', r['update_correct'])} "
            f"| {fmt('update_omission', r['update_omission'])} "
            f"| {fmt('qa_correct', r['qa_correct'])} "
            f"| {fmt('qa_hallucination', r['qa_hallucination'])} "
            f"| {fmt('qa_omission', r['qa_omission'])} |"
        )

    lines.append("")
    lines.append("> **Bold** = best value in each column.")
    lines.append("")

    with open(MD_PATH, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))
    print(f"✅ Markdown saved → {MD_PATH}")


# ── Entry point ───────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--open", action="store_true", help="Open Excel after update")
    args = parser.parse_args()

    runs = scan_runs()
    if not runs:
        print("⚠️  No completed runs found in results/")
        return

    build_excel(runs)
    build_markdown(runs)

    if args.open:
        subprocess.run(["open", MD_PATH])


if __name__ == "__main__":
    main()
